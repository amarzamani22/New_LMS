#----------------------------------------------------
# LMS Analysis Dashboard (Streamlit)
#----------------------------------------------------
# 1)This code loads QC workbooks for the LMS
# 2)Detects statistically significant outliers, and links them to Variance Report
# 3)Show list of FI that contribute to the detected outliers
#-----------------------------------------------------

import os
import re
from typing import Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl.utils import get_column_letter
from datetime import datetime
import openpyxl
import io

# ================ Page & Styles in CSS =================
st.set_page_config(page_title="LMS Analysis Dashboard", page_icon="🏦", layout="wide")
st.markdown("""""", unsafe_allow_html=True)

# ================ Constants / Configuration ================

# Mapping datasets name to specific QC sheet names in Excel workbook
SHEET_MAP: Dict[str, str] = {
    'Q1A: Employees': 'QC_Q1A_Main',
    'Q2A: Salary': 'QC_Q2A_Main',
    'Q3: Hours Worked': 'QC_Q3',
    'Q4: Vacancies': 'QC_Q4',
    'Q5: Separations': 'QC_Q5'
}
ALL_MONTHS: List[str] = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
ENTITY_COL = "Entity / Group"                 #Column name used to identify entities or groups in QC sheet
SUBQ_COL = "Subquestion"                      #Column name for subquestion (except for Question 3(Hours Worked))
WC_COL = "Worker Category"                    #Column name for worker category
ROLLUP_KEY = "All Financial Institutions"     #Roll-up label

# ---- Minimal fallback hierarchy (used only if no CSV is provided) ----
HIERARCHY_MAP: Dict[str, List[str]] = {
    "All Financial Institutions": ["Banking Institution", "DFI", "Insurans/Takaful Operators"],
    "Banking Institution": [
        "Commercial Banks", "Digital Banks", "Foreign Banks",
        "Islamic Banks", "Investment Banks", "International Islamic Banks"
    ],
    "DFI": [],  # Actual DFI list will come from CSV when available
    "Insurans/Takaful Operators": ["Insurers", "Takaful Operators"],
}

# ====================== Data Access (QC Workbook) ======================
@st.cache_data
def load_qc_sheet(year: int, sheet_name: str) -> pd.DataFrame | str:
    #Load a QC sheet for a give year and sheet name
    #Returns a DataFrame on success, or an error string on failure
    #Caches results to avoid re-reading repeatedly
    path = fr"C:\Users\ttamarz\OneDrive - Bank Negara Malaysia\RLMS\Output\QC Template\qc_workbook_{year}.xlsx"

    if not os.path.exists(path):
        return f"Error: File not found → {path}"
    try:
        # QC sheets have 5 header rows above the data header
        df = pd.read_excel(path, sheet_name=sheet_name, header=5)

        # Remove empty columns
        df.dropna(axis=1, how='all', inplace=True)

        # Normalize quarter total column names
        df.rename(columns={'Q1.1':'Q1_Total','Q2.1':'Q2_Total','Q3.1':'Q3_Total','Q4.1':'Q4_Total'}, inplace=True)
        return df
    except Exception as e:
        return f"Error reading sheet '{sheet_name}' from {path}: {e}"
    
# ============ Helpers for MHS =================
def mhs_page():
    """
    FINAL VERSION:
    User selects YEAR + QUARTER.
    System auto-fills:
        - 3 months
        - quarter row
        - year row if Q4
    Shows preview table (month breakdown + quarter summary).
    Prevents overwriting existing months.
    """
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter
    from copy import copy
    from datetime import datetime

    st.header("MHS Table Generator — Quarter Auto-Fill")

    # ---- USER INPUTS ----
    current_year = datetime.now().year
    years = list(range(current_year - 5, current_year + 2))
    year = st.selectbox("Select Year", years, index=years.index(current_year))
    quarter = st.selectbox("Select Quarter", ["Q1", "Q2", "Q3", "Q4"])

    quarter_map = {
        "Q1": [1, 2, 3],
        "Q2": [4, 5, 6],
        "Q3": [7, 8, 9],
        "Q4": [10, 11, 12],
    }
    q_months = quarter_map[quarter]

    mhs_file = st.file_uploader("Upload MHS Template (.xlsx)", type=["xlsx"])
    if not mhs_file:
        return

    # ---- TEMPLATE ANCHORS ----
    ANCHOR = {
        "year_block": 15,        # B15
        "quarter_block": 55,     # C55
        "month_block": 165       # C165
    }
    start_col = 5  # Column D

    # ---- LOAD TEMPLATE ----
    wb = openpyxl.load_workbook(mhs_file)
    ws = wb[wb.sheetnames[0]]

    # ---- HELPER: FIND NEXT EMPTY ROW ----
    def next_empty(sheet, start_row, col):
        r = start_row
        while sheet[f"{col}{r}"].value not in (None, ""):
            r += 1
        return r

    # ---- HELPER: CHECK IF MONTH ALREADY EXISTS ----
    def month_exists(sheet, y, m):
        r = ANCHOR["month_block"]
        while r < 5000:
            cell = sheet[f"C{r}"].value
            if cell == m:
                # check nearest year above
                for back in range(0, 10):
                    yc = sheet[f"B{r-back}"].value
                    if yc == y:
                        return True
            r += 1
        return False

    # ---- STOP IF ANY MONTH ALREADY EXISTS ----
    for m in q_months:
        if month_exists(ws, year, m):
            st.error(f"❌ Month {m} of {year} already exists in MHS table.")
            return

    # ---- LOAD QC SHEETS ----
    try:
        df_q1 = load_qc_sheet(year, SHEET_MAP["Q1A: Employees"])
        df_q4 = load_qc_sheet(year, SHEET_MAP["Q4: Vacancies"])
        df_q5 = load_qc_sheet(year, SHEET_MAP["Q5: Separations"])
    except Exception as e:
        st.error(f"Failed to load QC sheets: {e}")
        return

    # ---- SUBQUESTION LABELS ----
    SUB = {
            "emp_A": "A. Number of Employees",
            "emp_B1": "B(i). Malaysian Employees",
            "emp_B2": "B(ii). Non-Malaysian Employees",

            "VAC": "A. Number of Job Vacancies as at End of the Month",
            "NEWJOB": "Number of Job Vacancies Due to New Jobs Created During the Month",

            "HIRE": "New Hires and Recalls",

            "QUIT": "A. Quits and resignation (except retirement)",
            "LAYOFF": "B. Total Layoffs and Discharges",
            "OTHER": "C. Other Separation"
        }

    # HELPER: extract QC monthly
    def qc(df, sub_label, month_idx:int):
        """
        Returns the numeric sum for `sub_label` for the given month index (1..12).
        Logic:
          - Prefer rows where the Entity column indicates 'All' (All Financial Institution / All FI)
          - Prefer Worker Category rows that look like 'All', 'Total', 'All workers'
          - If no worker-total row exists, sum across available worker-category rows
        """
        if df is None or isinstance(df, str):
            return 0.0

        # helper to find entity column (common names)
        entity_col = None
        for cand in ["Entity / Group", "Entity", "Entity/Group", "Entity / group"]:
            if cand in df.columns:
                entity_col = cand
                break

        wc_col = None
        for cand in ["Worker Category", "Worker_Category", "WC", "Worker category"]:
            if cand in df.columns:
                wc_col = cand
                break

        month_col = ALL_MONTHS[month_idx - 1]  # "Jan".."Dec"
        # 1) exact subquestion match
        sel = df[df["Subquestion"].astype(str).str.strip() == sub_label] if "Subquestion" in df.columns else df.iloc[0:0]

        # 2) fallback: contains
        if sel.empty:
            sel = df[df["Subquestion"].astype(str).str.contains(sub_label.split()[0], na=False, case=False)]

        if sel.empty:
            # nothing matches
            return 0.0

        # 3) prefer rows where Entity indicates "All Financial Institution"
        if entity_col:
            all_entity_mask = sel[entity_col].astype(str).str.contains(r"all\s*financial|all\s*fi|all\s*institution|^all\b", case=False, na=False)
            if all_entity_mask.any():
                sel = sel[all_entity_mask]

        # 4) if Worker Category column present, prefer rows that are totals
        if wc_col and wc_col in sel.columns:
            wc_vals = sel[wc_col].astype(str).str.strip().fillna("")
            # patterns that likely indicate total row
            total_mask = wc_vals.str.contains(r"all|total|all workers|total workers", case=False, na=False)
            if total_mask.any():
                sel = sel[total_mask]
                # sum numeric month column across matching total rows
                try:
                    return float(pd.to_numeric(sel[month_col], errors="coerce").fillna(0).sum())
                except Exception:
                    return 0.0
            else:
                # no explicit total row — sum across available worker categories
                try:
                    return float(pd.to_numeric(sel[month_col], errors="coerce").fillna(0).sum())
                except Exception:
                    return 0.0
        else:
            # no worker category column — just sum the month column across sel
            try:
                return float(pd.to_numeric(sel[month_col], errors="coerce").fillna(0).sum())
            except Exception:
                return 0.0

    # ---- COLLECT VALUES FOR 3 MONTHS ----
    month_rows = []
    for m in q_months:
        # Employment: sum of A + B(i) + B(ii)
        emp_A = qc(df_q1, SUB["emp_A"], m)
        emp_B1 = qc(df_q1, SUB["emp_B1"], m)
        emp_B2 = qc(df_q1, SUB["emp_B2"], m)
        EMP = emp_A + emp_B1 + emp_B2

        VAC = qc(df_q4, SUB["VAC"], m)
        NEWJ = qc(df_q4, SUB["NEWJOB"], m)
        HIRE = qc(df_q5, SUB["HIRE"], m)

        QUIT = qc(df_q5, SUB["QUIT"], m)
        LAY = qc(df_q5, SUB["LAYOFF"], m)
        OTH = qc(df_q5, SUB["OTHER"], m)
        SEP = QUIT + LAY + OTH

        month_rows.append([
            m, EMP, VAC, NEWJ, HIRE, SEP, QUIT, LAY, OTH
        ])

    # ---- PREVIEW (MONTH BREAKDOWN) ----
    st.subheader("📅 Monthly Breakdown")
    df_prev_months = pd.DataFrame(month_rows, columns=[
        "Month", "Employment", "Vacancies", "New Jobs",
        "Hires", "Separations Total", "Quits", "Layoff", "Other"
    ])
    st.dataframe(df_prev_months)

    # ---- PREVIEW (QUARTER SUMMARY) ----
    st.subheader("📊 Quarter Summary")
    quarter_totals = df_prev_months.iloc[:, 1:].sum()
    st.dataframe(quarter_totals.to_frame("Value"))

    # ---- WRITE TO EXCEL ----
    if st.button("Write Quarter to MHS Template"):
        # Write 3 months
        for m, *vals in month_rows:
            r = next_empty(ws, ANCHOR["month_block"], "C")
            ws[f"B{r}"] = year
            ws[f"C{r}"] = m
            for i, v in enumerate(vals):
                ws[f"{get_column_letter(start_col + i)}{r}"] = v

        # Write quarter
        qr = next_empty(ws, ANCHOR["quarter_block"], "C")
        ws[f"B{qr}"] = year
        ws[f"C{qr}"] = quarter
        for i, v in enumerate(quarter_totals):
            ws[f"{get_column_letter(start_col + i)}{qr}"] = float(v)

        # If Q4 → write Year row
        if quarter == "Q4":
            yr_r = next_empty(ws, ANCHOR["year_block"], "B")
            year_totals = df_prev_months.iloc[:, 1:].sum()
            ws[f"B{yr_r}"] = year
            for i, v in enumerate(year_totals):
                ws[f"{get_column_letter(start_col + i)}{yr_r}"] = float(v)

        # Export
        out = io.BytesIO()
        wb.save(out)
        st.success("Quarter inserted successfully.")
        st.download_button(
            "Download Updated MHS File",
            data=out.getvalue(),
            file_name=f"MHS_{year}_{quarter}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@st.cache_data
def get_reporting_quarter(year: int) -> int:

    #Read _About sheet to know the current quarter in that workbook.
    path = fr"C:\Users\ttamarz\OneDrive - Bank Negara Malaysia\RLMS\Output\QC Template\qc_workbook_{year}.xlsx"
    try:
        about = pd.read_excel(path, sheet_name="_About", header=None)
        row = about[about[0] == "Quarter"]
        if not row.empty:
            # Extract the integer form something like "Q2" or "Quarter: Q3"
            qn = int(re.search(r'\d+', str(row.iloc[0,1])).group())
            return max(1, min(4, qn))
    except Exception:
        pass
    return 4 

def months_for_q(q: int) -> List[str]:
    # Return month labels for a given quarter number
    return {1:['Jan','Feb','Mar'], 2:['Apr','May','Jun'], 3:['Jul','Aug','Sep'], 4:['Oct','Nov','Dec']}[int(q)]

def current_q_month_labels(current_year: int, reporting_q: int) -> List[str]:
    #Return labels like '2025-Feb' for months in the current reporting quarter
    return [f"{current_year}-{m}" for m in months_for_q(reporting_q)]

def qc_row_slice(df: pd.DataFrame, entity: str, wc: str, subq: str) -> pd.DataFrame:
    #Return the row for a (Entity, Worker Category, Subquestion)
    cond = (df[ENTITY_COL] == entity) & (df[WC_COL] == wc)
    if SUBQ_COL in df.columns and subq != "N/A":
        cond &= (df[SUBQ_COL] == subq)
    return df[cond]

# ================== VR loader & matching helpers (to fetch FI justifications) =====================
def _norm(s) -> str:
    # Normalize strings: trim, collapse spaces, lowercase
    return re.sub(r"\s+", " ", str(s).strip().lower())

@st.cache_data
def load_vr_variance(vr_path: str) -> pd.DataFrame | str:
    # Load the VR (staging) Excel and prepare helper columns for matching
    # Returns DataFrame if available, 'PENDING' if path is blank/missing, or error message on failure
    if not vr_path or not os.path.exists(vr_path):
        return "PENDING"  # special marker to show 'Pending submission'
    try:
        df = pd.read_excel(vr_path, sheet_name="Variance")
        needed = ["Entity Name","Year","Quarter","Month","Question","Subquestion","Worker Category","%Growth","Justification"]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            return f"VR file missing columns: {missing}"
        
        # Normalize key matching columns
        df["_ent"] = df["Entity Name"].map(_norm)
        df["_subq"] = df["Subquestion"].map(_norm)
        df["_wc"] = df["Worker Category"].map(_norm)
        # Extract quarter number as numeric from a value like "Q1" "Q2"
        df["_qnum"] = df["Quarter"].astype(str).str.extract(r"\((\d)\)").astype(float)
        # Normalize month to Jan/Feb/... (first 3 letters)
        df["_month"]= df["Month"].astype(str).str[:3]
        return df
    except Exception as e:
        return f"Error reading VR file: {e}"

def _question_code_from_dataset(dataset_key: str) -> str:
    # Map Q1A: Employees -> Q1A for matching against VR's 'Question'
    return dataset_key.split(":")[0].replace(" ", "")

def find_vr_just_for_periods(
    vr_df: pd.DataFrame | str,
    dataset_key: str,
    entity_name: str,
    subq: str,
    wc: str,
    periods: List[str],   # e.g. ["2025-Feb"] or ["Q1 2025"]
) -> str:
    # Look up FI 'Justification' text in the VR file for the specified (entity/subquestion/worker category) and periods.
    # Accept either monthly labels "YYY-Mmm" or quarterly labels "Qx YYYY"
    # If multiple justifications match, deduplicate and join them with newlines.

    # Handle error / pending states cleanly
    if isinstance(vr_df, str):
        return "Pending submission" if vr_df == "PENDING" else vr_df
    
    qcode = _question_code_from_dataset(dataset_key)
    ent_norm = _norm(entity_name)
    subq_norm = _norm(subq)
    wc_norm = _norm(wc)
    justs = []

    for p in periods:
        # Monthly label "YYYY-Mmm"
        if "-" in p and "Q" not in p:
            yr_str, mon = p.split("-")
            try:
                yr = int(yr_str)
            except:
                continue
            # Strict month match
            sub = vr_df[
                (vr_df["Year"] == yr) &
                (vr_df["_month"] == mon) &
                (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                (vr_df["_ent"] == ent_norm) &
                (vr_df["_wc"] == wc_norm)
            ]
            if subq_norm and subq_norm != _norm("N/A"):
                sub = sub[sub["_subq"] == subq_norm]
            
            # Fallback: use quarter row if monthly row missing
            if sub.empty:
                q_from_mon = {"Jan":1,"Feb":1,"Mar":1,"Apr":2,"May":2,"Jun":2,
                              "Jul":3,"Aug":3,"Sep":3,"Oct":4,"Nov":4,"Dec":4}.get(mon, None)
                if q_from_mon is not None:
                    sub = vr_df[
                        (vr_df["Year"] == yr) &
                        (vr_df["_qnum"] == q_from_mon) &
                        (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                        (vr_df["_ent"] == ent_norm) &
                        (vr_df["_wc"] == wc_norm)
                    ]
                    if subq_norm and subq_norm != _norm("N/A"):
                        sub = sub[sub["_subq"] == subq_norm]
        else:
            # Quarterly label "Qx YYYY"
            try:
                qlab, yr_str = p.split()
                qn = int(qlab[1:])
                yr = int(yr_str)
            except:
                continue
            sub = vr_df[
                (vr_df["Year"] == yr) &
                (vr_df["_qnum"] == qn) &
                (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                (vr_df["_ent"] == ent_norm) &
                (vr_df["_wc"] == wc_norm)
            ]
            if subq_norm and subq_norm != _norm("N/A"):
                sub = sub[sub["_subq"] == subq_norm]

        # Collect non-empty justifications and dedupe
        js = [j for j in sub["Justification"].astype(str).tolist()
              if str(j).strip() and str(j).strip().lower() != "nan"]
        if js:
            justs.append(" \n".join(sorted(set(js))))

    if not justs:
        return "—"
    return " \n".join(sorted(set(justs)))

# ===================== Multi-year Series Builders (monthly/quarterly) ========================
def build_multi_year_monthly_series(
    entity: str, wc: str, subq: str, sheet_key: str,
    start_year: int, end_year: int, excluded_years: set[int]
) -> Tuple[pd.Series, Optional[pd.Series]]:
    
    # Build a monthly time series from start_year..end_year (exlcuding excluded_years)
    # Index like 'YYY-Mmm'
    # Also returns a "YoY alignment series" (previous year's same month values aligned on current index) for YoY% calculations 
    vals, idx = [], []

    for yr in range(int(start_year), int(end_year)+1):
        if yr in excluded_years: 
            continue

        df = load_qc_sheet(yr, SHEET_MAP[sheet_key])
        if isinstance(df, str): 
            continue

        row = qc_row_slice(df, entity, wc, subq)
        if row.empty: 
            continue

        rq = get_reporting_quarter(yr)
        months = [m for m in ALL_MONTHS[:rq*3] if m in row.columns]
        if not months: 
            continue

        s = pd.to_numeric(row[months].iloc[0], errors="coerce").astype(float)

        for m in months:
            idx.append(f"{yr}-{m}")
            vals.append(s.get(m, np.nan))

    series = pd.Series(vals, index=idx, dtype=float)

    # Build YoY alignment for each YYYY-Mmm, pick value from (YYYY-1)-Mmm if available
    yoy_vals, yoy_idx = [], []

    for lab in series.index:
        yr, mon = lab.split('-')
        prev = f"{int(yr)-1}-{mon}"
        if prev in series.index and not pd.isna(series[prev]):
            yoy_idx.append(lab); yoy_vals.append(series[prev])

    yoy_series = pd.Series(yoy_vals, index=yoy_idx, dtype=float) if yoy_idx else None
    return series, yoy_series

def build_multi_year_quarterly_series(
    entity: str, wc: str, subq: str, sheet_key: str,
    start_year: int, end_year: int, excluded_years: set[int]
) -> Tuple[pd.Series, Optional[pd.Series]]:
    # Build a quarterly time series. Uses Qx_Total if present; otherwise sums the months of each quarter
    # Index like 'Qx YYYY'
    # Also constructs a YoY alignment series for YoY% calculations
    labels, values = [], []

    for yr in range(int(start_year), int(end_year)+1):
        if yr in excluded_years: 
            continue

        df = load_qc_sheet(yr, SHEET_MAP[sheet_key])
        if isinstance(df, str): 
            continue

        row = qc_row_slice(df, entity, wc, subq)
        if row.empty: 
            continue

        rq = get_reporting_quarter(yr)
        
        for q in range(1, rq+1):
            col = f"Q{q}_Total"
            if col in row.columns and not pd.isna(row.iloc[0][col]):
                v = float(pd.to_numeric(row.iloc[0][col], errors="coerce"))
            else:
                # Sum months if Qx_Total is missing
                mlist = [m for m in months_for_q(q) if m in row.columns]
                v = float(pd.to_numeric(row[mlist].iloc[0], errors="coerce").astype(float).sum()) if mlist else np.nan

            if not np.isnan(v):
                labels.append(f"Q{q} {yr}")
                values.append(v)

    series = pd.Series(values, index=labels, dtype=float)

    # Build YoY alignment for quarters
    yoy_vals, yoy_idx = [], []

    for lab in series.index:
        q, yr_str = lab.split()
        prev = f"{q} {int(yr_str)-1}"
        if prev in series.index and not pd.isna(series[prev]):
            yoy_idx.append(lab); yoy_vals.append(series[prev])

    yoy_series = pd.Series(yoy_vals, index=yoy_idx, dtype=float) if yoy_idx else None
    return series, yoy_series

# ===================== Outlier detection engine ====================

def find_outliers_v2(
    series: pd.Series,
    yoy_series: Optional[pd.Series],
    pct_thresh: float,   # MoM/QoQ % threshold (gate)
    abs_cutoff: float,   # absolute change threshold (significance)
    iqr_k: float,        # IQR multiplier (significance)
    yoy_thresh: float    # YoY % threshold (significance)
) -> pd.DataFrame:
    
    # Outlier logic (two-stage):
    # 1) Gate: High MoM%/QoQ% if |MoM| >= pct_thresh AND |Δ| >= abs_cutoff
    # 2) Only then flag if (YoY >= yoy_thresh OR value is IQR outlier)
    # Reasons include: "High MoM%", "High YoY%", "IQR Detect Outlier"
    # Returns a DataFrame with index=Period and columns: Value, Statistical Reasons.
    
    out = []
    clean = series.dropna()
    if clean.size < 2:
        return pd.DataFrame(columns=["Period", "Value", "Statistical Reasons"]).set_index("Period")
    
    # IQR bounds
    q1, q3 = clean.quantile(0.25), clean.quantile(0.75)
    iqr = q3 - q1
    lb, ub = ((q1 - iqr_k * iqr, q3 + iqr_k * iqr) if iqr > 0 else (None, None))

    for i, (period, cur) in enumerate(series.items()):
        if pd.isna(cur) or i == 0: 
            continue
        prev = series.iloc[i - 1]
        if pd.isna(prev) or prev == 0: 
            continue

        abs_chg = cur - prev
        mom = abs_chg / prev

        # Stage 1: Gate on MoM/QOQ % AND absolute change
        if abs(mom) >= pct_thresh and abs(abs_chg) >= abs_cutoff:
            reasons = [f"High MoM% ({mom:+.0%})"]
            extra_reason = False

            # IQR anomaly on level
            if iqr > 0 and (cur < lb or cur > ub):
                reasons.append("IQR Detect Outlier")
                extra_reason = True

            # YoY significance if aligned previous year's period exists
            if yoy_series is not None and period in yoy_series.index:
                py = yoy_series.get(period)
                if not pd.isna(py) and py != 0:
                    yoy = (cur - py) / py
                    if abs(yoy) >= yoy_thresh:
                        reasons.append(f"High YoY% ({yoy:+.0%})")
                        extra_reason = True

            # Only flag if at least one significance reason exists
            if extra_reason:
                out.append({
                    "Period": period,
                    "Value": f"{cur:,.2f}",
                    "Statistical Reasons": ", ".join(reasons)
                })
    return pd.DataFrame(out).set_index("Period") if out else pd.DataFrame(columns=["Period", "Value", "Statistical Reasons"]).set_index("Period")

# ============= Charting: dual axis bar (levels) + line (% change) + outlier markers
def plot_dual_axis_with_outliers(
    series: pd.Series,
    growth_pct: pd.Series,
    outliers_focus: pd.DataFrame,
    title: str,
    left_title: str = "Value",
    right_title: str = "% Change (MoM)"
):
    # Render a Plotly Chart:
    # - Left axis (bars): level values
    # - Right axis (line): % change vs previous period
    # - Red X markers for 'true outliers' within the focus window

    x = list(series.index)
    y = series.values.astype(float)
    g = growth_pct.reindex(series.index).fillna(0)

    fig = go.Figure()

    # Bars for values (left axis)
    fig.add_trace(go.Bar(
        x=x, y=y, name=left_title, yaxis="y1", opacity=0.75,
        hovertemplate='Period: %{x}<br>Value: %{y:,.0f}'
    ))

    # Line for % change (right axis)
    fig.add_trace(go.Scatter(
        x=x, y=g.values, name=right_title, yaxis="y2",
        mode="lines+markers", line=dict(width=2, dash="dot", color="#003366"),
        hovertemplate='Period: %{x}<br>% Growth: %{y:+.0f}%'
    ))
    # Outlier markers (red X)
    if not outliers_focus.empty:
        ox = [p for p in outliers_focus.index if p in series.index]
        oy = [float(g.loc[p]) for p in ox]
        oreason = [outliers_focus.loc[p, "Statistical Reasons"] for p in ox]
        fig.add_trace(go.Scatter(
            x=ox, y=oy, mode='markers', name='True Outlier', yaxis="y2",
            marker=dict(symbol='x', size=14, color='red', line_width=2),
            hovertemplate='Period: %{x}<br>% Growth: %{y:+.0f}%<br>%{customdata}',
            customdata=oreason
        ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=16), x=0.5, xanchor='center'),
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0),
        margin=dict(l=10, r=10, t=70, b=10),
        xaxis=dict(tickangle=45),
        yaxis=dict(title=left_title, side='left'),
        yaxis2=dict(title=right_title, overlaying='y', side='right', showgrid=False, tickformat=".0f")
    )
    st.plotly_chart(fig, use_container_width=True)

# ==================== Hierarchy Loader (Parent-Child OR Institution/Sector/Subsector)
@st.cache_data
def load_hierarchy_any(path: str) -> Dict[str, List[str]]:
    
    # Accepts either:
    #  - Parent,Child CSV (used directly), or
    #  - Institution,Sector,Subsector CSV (auto-builds AFI tree):
    #       AFI -> Banking Institution / DFI / Insurans-Takaful
    #       Banking Institution -> (Commercial, Investment, Foreign, Islamic, Digital, IIB) -> entities
    #       DFI -> entities directly
    #       Insurans/Takaful -> (Insurers, Takaful Operators) -> entities
    # Removes self-child edges.

    import pandas as pd, os
    if not path or not os.path.exists(path):
        return {}

    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # Case A: Parent-Child edges provided directly
    if {"Parent","Child"} <= set(df.columns):
        df["Parent"] = df["Parent"].astype(str).str.strip()
        df["Child"]  = df["Child"].astype(str).str.strip()
        df = df[df["Parent"].str.lower() != df["Child"].str.lower()]
        return df.groupby("Parent")["Child"].apply(list).to_dict()

    # Case B: Institution,Sector,Subsector (build All FI tree)
    if {"Institution","Sector","Subsector"} <= set(df.columns):
        AFI  = "All Financial Institutions"
        BI   = "Banking Institution"
        DFI  = "DFI"
        ITOP = "Insurans/Takaful Operators"

        df["Institution"] = df["Institution"].astype(str).str.strip()
        df["Sector"]      = df["Sector"].astype(str).str.strip()
        df["Subsector"]   = df["Subsector"].astype(str).str.strip()

        banking_subs = {"Commercial Banks","Investment Banks","Foreign Banks","Islamic Banks","Digital Banks","International Islamic Banks"}
        ins_tak_subs = {"Insurers","Takaful Operators"}
        present_subs = set(df["Subsector"].unique())

        edges = []

        # All FI layer
        if (banking_subs & present_subs): edges.append({"Parent": AFI, "Child": BI})
        if ("DFI" in present_subs):      edges.append({"Parent": AFI, "Child": DFI})
        if (ins_tak_subs & present_subs):edges.append({"Parent": AFI, "Child": ITOP})

        # Banking branches
        for s in sorted(banking_subs & present_subs):
            edges.append({"Parent": BI, "Child": s})
        # Ins/Takaful branches
        for s in sorted(ins_tak_subs & present_subs):
            edges.append({"Parent": ITOP, "Child": s})

        # DFI -> leaf institutions (NO subsector under DFI)
        if "DFI" in present_subs:
            for inst in sorted(df.loc[df["Subsector"] == "DFI", "Institution"].unique()):
                edges.append({"Parent": DFI, "Child": inst})

        # Banking leaves
        for s in sorted(banking_subs & present_subs):
            for inst in sorted(df.loc[df["Subsector"] == s, "Institution"].unique()):
                edges.append({"Parent": s, "Child": inst})

        # Ins/Takaful leaves
        for s in sorted(ins_tak_subs & present_subs):
            for inst in sorted(df.loc[df["Subsector"] == s, "Institution"].unique()):
                edges.append({"Parent": s, "Child": inst})

        hier_df = pd.DataFrame(edges).drop_duplicates()
        hier_df = hier_df[hier_df["Parent"].str.lower() != hier_df["Child"].str.lower()]
        return hier_df.groupby("Parent")["Child"].apply(list).to_dict()

    # If neither schema matches, warn and return empty
    st.warning("Hierarchy CSV must have either (Parent,Child) or (Institution,Sector,Subsector) columns.")
    return {}

# ================ Attribution helpers: discover leaf entities present in QC =======================

def _norm2(s: str) -> str:
    # Normalize but keep casefold (better for non-english) and preserve space a single spaces
    return re.sub(r"\s+", " ", str(s).strip()).casefold()

def _canonical_map(series: pd.Series) -> Dict[str, str]:
    # Build a map 'normalized name' -> 'most frequent display variant' observed in QC.
    # This avoids display mismatches when the same entity appears with minor naming differences.
    tmp = series.dropna().astype(str)
    norm = tmp.map(_norm2)
    canon_map = (
        pd.DataFrame({"orig": tmp, "norm": norm})
        .groupby("norm")["orig"]
        .agg(lambda s: s.value_counts().idxmax())
        .to_dict()
    )
    return canon_map

def get_children_in_data(df: pd.DataFrame, parent: str, hier: Dict[str, List[str]]) -> List[str]:
    
    # Return immediate children of 'parent' that actually appear in QC (or exist as groups to recurse into).
    # - Ignores self-child edges.
    # - If a child group exists in the hierarchy but has no direct QC row, we keep it to recurse down later.
    if parent not in hier:
        return []
    
    candidates = [c for c in hier[parent] if _norm2(c) != _norm2(parent)]

    # Build canonical map for QC entities to ensure display names match QC variants
    canon = _canonical_map(df[ENTITY_COL])
    present_norms = set(canon.keys())

    matched = []
    for c in candidates:
        n = _norm2(c)
        if n in present_norms:
            matched.append(canon[n])     # exact QC display name
        elif c in hier:
            matched.append(c)            # group exists in hierarchy (no direct QC row)
    return list(dict.fromkeys(matched))

def flatten_leaves_in_qc(df: pd.DataFrame, root: str, hier: Dict[str, List[str]]) -> List[str]:
    
    # Recursively collect leaf entities under 'root' that actually exist in QC.
    # A leaf is a child that is not a parent in 'hier' (i.e., no further children).
    
    leaves = set()

    def _recurse(node: str):
        kids = get_children_in_data(df, node, hier)
        if not kids:
            #If node is a real entity row in QC, mark as leaf
            if node in df[ENTITY_COL].unique():
                leaves.add(node)
            return
        for k in kids:
            if k in hier:
                _recurse(k)
            else:
                leaves.add(k)

    _recurse(root)

    # Remove known group labels if they slipped in
    groups_to_exclude = {
        "All Financial Institutions", "Banking Institution", "DFI",
        "Insurans/Takaful Operators", "Commercial Banks", "Investment Banks",
        "Foreign Banks", "Islamic Banks", "Digital Banks", "International Islamic Banks",
        "Insurers", "Takaful Operators"
    }
    return sorted([e for e in leaves if e not in groups_to_exclude])

def compute_entity_contributions(
    entities: List[str],
    subq: str,
    wc: str,
    sheet_key: str,
    start_year: int,
    end_year: int,
    exclude_years: set[int],
    period_label: str,     # "YYYY-Mmm" or "Qx YYYY"
    time_view: str         # "Monthly" or "Quarterly"
) -> pd.DataFrame:
    
    # For a chosen outlier period, compute per-entity:
    #  - Prev (previous period value), Curr (current period), Δ (Curr-Prev)
    #  - Contribution % relative to the total Δ across entities
    # Sorted by |Δ| descending to surface top movers.

    rows = []

    for ent in entities:
        if time_view == "Monthly":
            s, _ = build_multi_year_monthly_series(ent, wc, subq, sheet_key, start_year, end_year, exclude_years)
        else:
            s, _ = build_multi_year_quarterly_series(ent, wc, subq, sheet_key, start_year, end_year, exclude_years)

        if s.empty or period_label not in s.index:
            continue
        idx_list = list(s.index)
        i = idx_list.index(period_label)
        if i == 0:
            continue  # no previous period
        prev, cur = float(s.iloc[i-1]), float(s.iloc[i])
        if np.isnan(prev) or np.isnan(cur):
            continue
        rows.append({"Entity": ent, "Prev": prev, "Curr": cur, "Delta": cur - prev})

    dfe = pd.DataFrame(rows)
    if dfe.empty:
        return dfe

    total_delta = dfe["Delta"].sum()
    dfe["Contribution %"] = np.where(total_delta != 0, dfe["Delta"] / total_delta * 100.0, np.nan)

    # Sort by magnitude of movement
    dfe.sort_values(by="Delta", key=np.abs, ascending=False, inplace=True)  # rank biggest movers
    dfe.reset_index(drop=True, inplace=True)
    return dfe

def plot_top_contributors_bar(dfe: pd.DataFrame, title: str = "Top Contributors (Δ vs previous)"):
    # Horizontal bar chart highlighting which entities contributed most 
    # Positive = increase, Negative = decrease
    if dfe.empty:
        st.info("No entity-level data available.")
        return
    
    show = dfe.copy()
    show["sign"] = np.where(show["Delta"] >= 0, "Increase", "Decrease")
    colors = {"Increase": "#2ca02c", "Decrease": "#d62728"}

    fig = go.Figure(go.Bar(
        x=show["Delta"],
        y=show["Entity"],
        orientation="h",
        marker=dict(color=[colors[s] for s in show["sign"]]),
        hovertemplate="Entity: %{y}<br>Δ: %{x:+,.0f}<extra></extra>"
    ))

    fig.update_layout(
        title=title,
        xaxis_title="Δ vs previous period",
        yaxis_title=None,
        margin=dict(l=10, r=10, t=50, b=10)
    )

    st.plotly_chart(fig, use_container_width=True)


# ==================== Sidebar UI Controls ==================

def sidebar_controls():
    st.sidebar.title("Analysis Controls")

    # Years / timeline selection
    years_available = list(range(2019, 2031))
    current_year = st.sidebar.selectbox(
        "Current Year (detect current quarter from this workbook):",
        options=sorted(years_available, reverse=True),
        index=years_available.index(2025)
    )
    start_year = st.sidebar.selectbox(
        "Start Year (timeline):",
        options=years_available,
        index=years_available.index(2022)
    )
    end_year = st.sidebar.selectbox(
        "End Year (timeline):",
        options=years_available,
        index=years_available.index(2025)
    )
    if end_year < start_year:
        st.sidebar.error("End Year must be ≥ Start Year.")
    exclude_years = st.sidebar.multiselect(
        "Exclude Years (optional):",
        options=[y for y in range(start_year, end_year+1)],
        default=[]
    )

    # Threshold for outlier detection
    st.sidebar.markdown("---")
    st.sidebar.subheader("Thresholds")
    mom_pct = st.sidebar.slider("MoM/QoQ % Threshold (Gate)", 0, 100, 25, 5, format="%d%%") / 100.0
    abs_cut = st.sidebar.slider("Absolute Change (Significance)", 10, 1000, 50, 10)
    iqr_k = st.sidebar.slider("IQR Sensitivity (Significance)", 1.0, 3.0, 1.5, 0.1)
    yoy_pct = st.sidebar.slider("YoY % Threshold (Significance)", 0, 100, 30, 5, format="%d%%") / 100.0

    # Frequency & dataset
    st.sidebar.markdown("---")
    time_view = st.sidebar.radio("Frequency:", options=['Monthly','Quarterly'], horizontal=True)
    dataset = st.sidebar.selectbox("Dataset:", options=list(SHEET_MAP.keys()))

    # Outlier focus window selection
    st.sidebar.markdown("---")
    st.sidebar.subheader("Outlier Focus")
    focus_mode = st.sidebar.radio(
        "Show outliers for:",
        options=["Current quarter", "Pick year & quarter"],
        index=0, horizontal=False
    )
    focus_year = None
    focus_quarter = None
    if focus_mode == "Pick year & quarter":
        focus_year = st.sidebar.selectbox(
            "Focus Year:",
            options=list(range(start_year, end_year + 1)),
            index=list(range(start_year, end_year + 1)).index(min(end_year, int(current_year)))
        )
        focus_quarter = st.sidebar.selectbox("Focus Quarter:", options=[1, 2, 3, 4], index=0)

    # VR staging file path
    st.sidebar.markdown("---")
    st.sidebar.subheader("VR Staging")
    vr_path = st.sidebar.text_input(
        "Full path to VR staging Excel (sheet 'Variance'):",
        value="",
        placeholder=r"C:\...\VR_Consol_2025_Quarter1.xlsx"
    )

    # Hierarchy CSV path 
    st.sidebar.markdown("---")
    st.sidebar.subheader("Hierarchy (optional)")
    hier_path = st.sidebar.text_input(
        "Path to Entity Hierarchy CSV (Parent,Child or Institution,Sector,Subsector):",
        value="",
        placeholder=r"C:\...\entity_hierarchy_full.csv"
    )

    return {
        "current_year": int(current_year),
        "start_year": int(start_year),
        "end_year": int(end_year),
        "exclude_years": set(map(int, exclude_years)),
        "mom_pct": float(mom_pct),
        "abs_cut": float(abs_cut),
        "iqr_k": float(iqr_k),
        "yoy_pct": float(yoy_pct),
        "time_view": time_view,
        "dataset": dataset,
        "focus_mode": focus_mode,
        "focus_year": int(focus_year) if focus_year is not None else None,
        "focus_quarter": int(focus_quarter) if focus_quarter is not None else None,
        "vr_path": vr_path.strip(),
        "hier_path": hier_path.strip(),
    }

# ==================== Main Interactive View ==================

def _is_total_wc(name: str) -> bool:
    # Consider anything that contains 'total' (case-insensitive) as a Total WC
    return "total" in str(name).casefold()

# ==========================================
# VR helper: Build VR table by Worker Category (non-destructive addition)
# ==========================================
def build_vr_wc_table_for_entity(vr_df: pd.DataFrame, df_cur: pd.DataFrame, dataset_key: str, entity: str, subq: str, periods_list: List[str]) -> pd.DataFrame:
    """
    Returns a dataframe of VR justifications by Worker Category for the given entity/subquestion and period list.
    If vr_df is "PENDING" or an error string, returns a single-row dataframe.
    """
    if isinstance(vr_df, str):
        return pd.DataFrame([{"Worker Category": "All workers", "FI Justification (selected period)": ("Pending submission" if vr_df == "PENDING" else vr_df)}])

    try:
        mask_ent = (df_cur[ENTITY_COL] == entity)
        if SUBQ_COL in df_cur.columns:
            mask_ent &= (df_cur[SUBQ_COL] == subq)
        wc_list = sorted(df_cur.loc[mask_ent, WC_COL].dropna().unique().tolist())
        if not wc_list:
            wc_list = ["All workers"]
    except Exception:
        wc_list = ["All workers"]

    rows = []
    for wc_any in wc_list:
        try:
            just_text = find_vr_just_for_periods(
                vr_df=vr_df,
                dataset_key=dataset_key,
                entity_name=entity,
                subq=subq,
                wc=wc_any,
                periods=periods_list
            )
        except Exception as e:
            just_text = f"(error reading VR: {e})"
        rows.append({"Worker Category": wc_any, "FI Justification (selected period)": just_text})

    show_wc = pd.DataFrame(rows)
    if not show_wc.empty:
        # Sort with "All workers" (or total-like) first, then alphabetical
        show_wc["__is_total"] = show_wc["Worker Category"].map(lambda x: 0 if _is_total_wc(str(x)) else 1)
        show_wc.sort_values(["__is_total", "Worker Category"], inplace=True, kind="stable")
        show_wc.drop(columns="__is_total", inplace=True)
    return show_wc


def main_view():
    st.title("🏦 LMS Analysis Dashboard")

    # Read sidebar configuration and show a summary 
    cfg = sidebar_controls()
    rq = get_reporting_quarter(cfg["current_year"])
    st.sidebar.info(f"Analyzing **{cfg['start_year']}–{cfg['end_year']}** (excl: {', '.join(map(str,cfg['exclude_years'])) or 'none'}) • Current workbook: **{cfg['current_year']} Q{rq}**")

    # Decide which period's outliers to highlight (current quarter or user-picked)
    if cfg["focus_mode"] == "Current quarter":
        focus_q = rq
        focus_year = cfg["current_year"]
    else:
        focus_q = cfg["focus_quarter"] or rq
        focus_year = cfg["focus_year"] or cfg["current_year"]

    focus_month_labels = set([f"{focus_year}-{m}" for m in months_for_q(focus_q)])
    focus_quarter_label = f"Q{focus_q} {focus_year}"

    # Badge display
    month_str = "–".join(months_for_q(focus_q))
    st.markdown(
        f'Outlier focus: **Q{focus_q} {focus_year}** ({month_str})',
        unsafe_allow_html=True
    )

    # Load the current-year QC sheet for the choosen dataset
    df_cur = load_qc_sheet(cfg["current_year"], SHEET_MAP[cfg["dataset"]])
    if isinstance(df_cur, str):
        st.error(df_cur)
        return
    


    # Normalize display variations for Entity/Subquestion/Worker Category to keep UI clean
    def _norm_local(s: str) -> str:
        return re.sub(r"\s+", " ", str(s).strip()).casefold()
    
    def _canonical(series: pd.Series) -> pd.Series:

        # Convert each value to its most frequent 'display' variant based on its normalized form
        tmp = series.dropna().astype(str)
        norm = tmp.map(_norm_local)
        canon_map = (
            pd.DataFrame({"orig": tmp, "norm": norm})
            .groupby("norm")["orig"]
            .agg(lambda s: s.value_counts().idxmax())
        )
        return norm.map(canon_map)

    df_cur["_ent_norm"] = df_cur[ENTITY_COL].map(_norm_local)
    df_cur["_ent_disp"] = _canonical(df_cur[ENTITY_COL])
    if SUBQ_COL in df_cur.columns:
        df_cur["_subq_norm"] = df_cur[SUBQ_COL].map(_norm_local)
        df_cur["_subq_disp"] = _canonical(df_cur[SUBQ_COL])
    df_cur["_wc_norm"] = df_cur[WC_COL].map(_norm_local)
    df_cur["_wc_disp"] = _canonical(df_cur[WC_COL])

    # Entity/Subquestion/Worker Category selection widget
    st.header(f"Outlier Detection: {cfg['dataset']}")
    entity = st.selectbox(
        "Entity / Group:",
        options=df_cur[ENTITY_COL].unique(),
        index=list(df_cur[ENTITY_COL].unique()).index(ROLLUP_KEY) if ROLLUP_KEY in df_cur[ENTITY_COL].unique() else 0
    )
    if SUBQ_COL in df_cur.columns and df_cur[df_cur[ENTITY_COL]==entity][SUBQ_COL].nunique() > 1:
        u_subq = df_cur[df_cur[ENTITY_COL]==entity][SUBQ_COL].unique()
        subq_default = np.where(u_subq == 'Employment = A+B(i)+B(ii)')[0][0] if 'Employment = A+B(i)+B(ii)' in u_subq else 0
        subq = st.selectbox("Subquestion:", options=u_subq, index=int(subq_default))
        u_wc = df_cur[(df_cur[ENTITY_COL]==entity) & (df_cur[SUBQ_COL]==subq)][WC_COL].unique()
        wc_default = np.where(u_wc == 'Total Employment')[0][0] if 'Total Employment' in u_wc else 0
        wc = st.selectbox("Worker Category:", options=u_wc, index=int(wc_default))
    else:
        subq = "N/A"
        u_wc = df_cur[df_cur[ENTITY_COL]==entity][WC_COL].unique()
        wc_default = np.where(u_wc == 'Total Employment')[0][0] if 'Total Employment' in u_wc else 0
        wc = st.selectbox("Worker Category:", options=u_wc, index=int(wc_default))

    st.caption(f"Displaying: {entity}  \n{subq}  \n{wc}")

    # Load VR staging once
    # String indicates "PENDING" or error
    vr_df = load_vr_variance(cfg["vr_path"])

    # Build the selected frequency series and run detection
    if cfg["time_view"] == "Monthly":
        series, yoy_series = build_multi_year_monthly_series(
            entity, wc, subq, cfg["dataset"],
            cfg["start_year"], cfg["end_year"],
            cfg["exclude_years"]
        )
        if series.empty or series.isnull().all():
            st.warning("No data found for the selected timeline/filters.")
            return
        
        # Detection outliers across the whole timeline
        out_all = find_outliers_v2(series, yoy_series, cfg["mom_pct"], cfg["abs_cut"], cfg["iqr_k"], cfg["yoy_pct"])

        # Keep only outliers within focused quarter's months
        out_focus = out_all.loc[out_all.index.intersection(focus_month_labels)]

        # Attach VR justifications per outlier period
        if not out_focus.empty:
            periods_list = out_focus.index.tolist()
            per_row_just = []
            if isinstance(vr_df, str):
                per_row_just = [("Pending submission" if vr_df == "PENDING" else vr_df)] * len(out_focus)
            else:
                for p in periods_list:
                    per_row_just.append(find_vr_just_for_periods(vr_df, cfg["dataset"], entity, subq, wc, [p]))

            out_focus = out_focus.copy()
            out_focus["FI Justification"] = per_row_just

        # %MoM growth for plotting (rounded to match VR)
        growth_pct = series.pct_change()
        growth_pct = growth_pct.replace([np.inf, -np.inf], np.nan)
        growth_pct = (growth_pct * 100).round().fillna(0)



        # Plot monthly chart
        plot_dual_axis_with_outliers(
            series=series, growth_pct=growth_pct, outliers_focus=out_focus,
            title=f"Monthly Trend ({cfg['start_year']}–{cfg['end_year']})"
        )

    else:
        # Quarterly path
        series, yoy_series = build_multi_year_quarterly_series(
            entity, wc, subq, cfg["dataset"],
            cfg["start_year"], cfg["end_year"],
            cfg["exclude_years"]
        )
        if series.empty or series.isnull().all():
            st.warning("No data found for the selected timeline/filters.")
            return
        
        # Detect outliers across the whole timeline
        out_all = find_outliers_v2(series, yoy_series, cfg["mom_pct"], cfg["abs_cut"], cfg["iqr_k"], cfg["yoy_pct"])

        # Keep only the focused quarter outlier
        out_focus = out_all.loc[out_all.index.intersection({focus_quarter_label})]

        # Attach VR justifications combining all months in that quarter
        if not out_focus.empty:
            qy = int(focus_quarter_label.split()[1])
            qn = int(focus_quarter_label.split()[0][1:])
            q_months = [f"{qy}-{m}" for m in months_for_q(qn)]
            per_row_just = []
            for _ in out_focus.index.tolist():  # usually one row
                per_row_just.append(
                    find_vr_just_for_periods(
                        vr_df=vr_df, dataset_key=cfg["dataset"], entity_name=entity,
                        subq=subq, wc=wc, periods=q_months
                    )
                )
            out_focus = out_focus.copy()
            out_focus["FI Justification"] = per_row_just

        # %QoQ growth for plotting
        growth_pct = series.pct_change()
        growth_pct = growth_pct.replace([np.inf, -np.inf], np.nan)
        growth_pct = (growth_pct * 100).round().fillna(0)

        # Plot quarterly chart
        plot_dual_axis_with_outliers(
            series=series, growth_pct=growth_pct, outliers_focus=out_focus,
            title=f"Quarterly Trend ({cfg['start_year']}–{cfg['end_year']})",
            right_title="% Change (QoQ)"
        )

    # KPI cards + outlier table for current focus
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Latest Value", f"{series.iloc[-1]:,.0f}")
    c2.metric("Average", f"{series.mean():,.0f}")
    c3.metric("Highest Value", f"{np.nanmax(series):,.0f}")
    c4.metric("Current-Q Outliers", len(out_focus))
    if not out_focus.empty:
        st.error("🚨 True Outlier(s) in Current Quarter")
        st.dataframe(out_focus, use_container_width=True)
    else:
        st.success("✅ No current-quarter outliers")

    # Attribution: which leaf entities contributed to the outlier?
    # Only shown when (a) there is at least one outlier in focus
    # (b) selected entity is the root roll-up "All Financial Institution"
    hier = load_hierarchy_any(cfg.get("hier_path","")) or HIERARCHY_MAP

    if not out_focus.empty and entity == ROLLUP_KEY:
        with st.expander("🔎 Contribution by FI", expanded=True):
            # Let the user pick which flagged period (within focus) to attribute
            attrib_period = st.selectbox(
                "Outlier period:",
                options=list(out_focus.index),
                index=0
            )

            # Resolve all leaves under All FI that actually exist in QC
            leaf_entities = flatten_leaves_in_qc(df_cur, ROLLUP_KEY, hier)
            if not leaf_entities:
                st.warning("No leaf entities found under All FI. Check the hierarchy path or mapping.")
            else:
                # Calculate contributions (Prev, Curr, Delta, Contributions %) per leaf for that period
                dfe = compute_entity_contributions(
                    entities=leaf_entities,
                    subq=subq,
                    wc=wc,
                    sheet_key=cfg["dataset"],
                    start_year=cfg["start_year"],
                    end_year=cfg["end_year"],
                    exclude_years=cfg["exclude_years"],
                    period_label=attrib_period,
                    time_view=cfg["time_view"]
                )

                if dfe.empty:
                    st.info("No entity had usable values in the selected period (or previous period).")
                else:
                    # Quick filtering and top-N selection
                    col_s, col_n = st.columns([2,1])
                    search_text = col_s.text_input("Filter entities (optional):", value="", placeholder="Type part of a name, e.g., 'MAYBANK'")
                    top_n = col_n.slider("Top N by |Δ|", min_value=5, max_value=50, value=15, step=5)

                    dfe_view = dfe.copy()
                    if search_text.strip():
                        q = search_text.strip().lower()
                        dfe_view = dfe_view[dfe_view["Entity"].str.lower().str.contains(q)]
                    dfe_view = dfe_view.head(top_n)

                    # Attach VR justification per entity for the same period
                    if isinstance(vr_df, str):
                        dfe_view["FI Justification"] = ("Pending submission" if vr_df == "PENDING" else vr_df)
                    else:
                        j_list = []
                        for ent in dfe_view["Entity"]:
                            j_list.append(
                                find_vr_just_for_periods(
                                    vr_df=vr_df,
                                    dataset_key=cfg["dataset"],
                                    entity_name=ent,
                                    subq=subq,
                                    wc=wc,
                                    periods=[attrib_period]
                                )
                            )
                        dfe_view["FI Justification"] = j_list



                    # --- Add-on: Interactive "FI Justification" click-to-expand (by Worker Category)
                    try:
                        show2 = dfe_view.copy()
                        show2["Prev"] = show2["Prev"].map(lambda v: f"{v:,.0f}")
                        show2["Curr"] = show2["Curr"].map(lambda v: f"{v:,.0f}")
                        show2["Delta"] = show2["Delta"].map(lambda v: f"{v:+,.0f}")
                        show2["Contribution %"] = show2["Contribution %"].map(lambda p: f"{p:+.1f}%" if pd.notna(p) else "–")
                        # Render a data editor with a clickable checkbox column
                        show2["FI Justification (all WC)"] = False

                        edited = st.data_editor(
                            show2,
                            use_container_width=True,
                            disabled=["Prev","Curr","Delta","Contribution %","FI Justification"],
                            column_config={
                                "FI Justification (all WC)": st.column_config.CheckboxColumn(
                                    "FI Justification", help="Click to view all Worker Categories for this FI"
                                )
                            },
                            key=f"fi_contrib_allwc_{attrib_period}"
                        )

                        # For any rows where the user ticked the FI Justification cell, render a breakdown by Worker Category
                        sel = edited[edited["FI Justification (all WC)"] == True]
                        if not sel.empty:
                            st.markdown("### VR Justifications by Worker Category (Selected FI)")
                            for _, _row in sel.iterrows():
                                _ent = _row["Entity"]
                                st.markdown(f"**{_ent} — {attrib_period}**")
                                wc_table = build_vr_wc_table_for_entity(
                                    vr_df=vr_df,
                                    df_cur=df_cur,
                                    dataset_key=cfg["dataset"],
                                    entity=_ent,
                                    subq=subq,
                                    periods_list=[attrib_period]
                                )
                                st.dataframe(wc_table, use_container_width=True)
                    except Exception as _e:
                        st.caption(f"(FYI: could not render interactive VR-by-worker-category view: {_e})")


                    # Visualize contributions
                    plot_top_contributors_bar(dfe_view, title=f"Top Contributors (All FI — {attrib_period})")
    else:
        # Keep UI minimal if not applicable
        pass

# ----------------------------
# Page Navigation (Added)
# ----------------------------
def router():
    page = st.sidebar.selectbox(
        "Select Page",
        ["Dashboard", "MHS Table"]
    )
    if page == "Dashboard":
        main_view()
    elif page == "MHS Table":
        mhs_page()
# ============ Run ================
if __name__ == "__main__":
    router()


