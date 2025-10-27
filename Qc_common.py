# qc_common.py
from __future__ import annotations
from dataclasses import dataclass
from datetime import datetime
from typing import Dict, List, Optional, Sequence, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import FormulaRule

# ================
# Quarter / months
# ================
QMAP = {
    "Quarter 1": "Q1", "Quarter 2": "Q2", "Quarter 3": "Q3", "Quarter 4": "Q4",
    "Q1": "Q1", "Q2": "Q2", "Q3": "Q3", "Q4": "Q4",
}
Q_TO_MONTHS = {
    "Q1": ["Jan","Feb","Mar"],
    "Q2": ["Apr","May","Jun"],
    "Q3": ["Jul","Aug","Sep"],
    "Q4": ["Oct","Nov","Dec"],
}
MONTHS_FULL = ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]

def normalize_quarter_label(qseries: pd.Series) -> str:
    if qseries is None or qseries.empty:
        return "Q1"
    tmp = qseries.astype(str).map(lambda s: QMAP.get(s.strip(), s.strip()))
    rank = {"Q1":1,"Q2":2,"Q3":3,"Q4":4}
    idx = tmp.map(rank).fillna(0).idxmax()
    return tmp.loc[idx]

def months_up_to(qlabel: str) -> List[str]:
    out: List[str] = []
    for q in ["Q1","Q2","Q3","Q4"]:
        out += Q_TO_MONTHS[q]
        if q == QMAP.get(str(qlabel).strip(), str(qlabel).strip()):
            break
    return out

# =========================
# Entity types & rollups
# =========================
ROLLUPS: Dict[str, List[str]] = {
    "All Financial Institutions": [
        "Commercial Banks", "Investment Banks", "Islamic Banks", "DFI",
        "Insurers", "Takaful Operators", "Foreign Banks",
        "International Islamic Banks", "Digital Banks"
    ],
    "Banking Institutions": [
        "Commercial Banks", "Investment Banks", "Islamic Banks",
        "Digital Banks", "Foreign Banks"
    ],
    "Development Financial Institution": ["DFI"],

    "Insurans/Takaful": ["Insurers", "Takaful Operators"],

    "Commercial Banks": ["Commercial Banks", "Foreign Banks"],
    "Investment Banks": ["Investment Banks"],
    "Islamic Banks": ["Islamic Banks"],
    "Digital Banks": ["Digital Banks"],
    "International Islamic Banks": ["International Islamic Banks"],
    "Foreign Banks": ["Foreign Banks"],
    "DFI": ["DFI"],
    "Insurers": ["Insurers"],
    "Takaful Operators": ["Takaful Operators"],
}
ROLLUP_ORDER = [
    "All Financial Institutions","Banking Institutions","Commercial Banks","Investment Banks",
    "Islamic Banks","Foreign Banks","Digital Banks","International Islamic Banks","DFI","Insurers","Takaful Operators",
]

# =============== Embedded FI → Type mapping (from your list) ===============
ENTITY_TO_TYPE: Dict[str, str] = {
    # (Banks/DFIs/Foreign/Islamic/Digital/IIB)
    "AFFIN BANK BERHAD":"Commercial Banks","AFFIN HWANG INVESTMENT BANK BERHAD":"Investment Banks",
    "AFFIN ISLAMIC BANK BERHAD":"Islamic Banks","AGROBANK (BANK PERTANIAN MALAYSIA BERHAD)":"DFI",
    "AL RAJHI BANKING & INVESTMENT CORPORATION (MALAYSIA) BHD":"Islamic Banks",
    "ALKHAIR INTERNATIONAL ISLAMIC BANK BHD":"International Islamic Banks",
    "ALLIANCE BANK MALAYSIA BERHAD":"Commercial Banks","ALLIANCE INVESTMENT BANK BERHAD":"Investment Banks",
    "ALLIANCE ISLAMIC BANK BERHAD":"Islamic Banks","AMBANK (M) BERHAD":"Commercial Banks",
    "AMBANK ISLAMIC BERHAD":"Islamic Banks","AMINVESTMENT BANK BERHAD":"Investment Banks",
    "MBSB BANK BERHAD":"Islamic Banks","BANGKOK BANK BERHAD":"Foreign Banks",
    "BANK ISLAM MALAYSIA BERHAD":"Islamic Banks","BANK KERJASAMA RAKYAT MALAYSIA BERHAD":"Islamic Banks",
    "BANK MUAMALAT MALAYSIA BERHAD":"Islamic Banks","BANK OF AMERICA MALAYSIA BERHAD":"Foreign Banks",
    "BANK OF CHINA (MALAYSIA) BERHAD":"Foreign Banks","BANK OF TOKYO-MITSUBISHI UFJ (MALAYSIA) BERHAD":"Foreign Banks",
    "BANK PEMBANGUNAN MALAYSIA BERHAD":"DFI","BANK SIMPANAN NASIONAL":"DFI",
    "BNP PARIBAS MALAYSIA BERHAD":"Foreign Banks","CHINA CONSTRUCTION BANK MALAYSIA BERHAD":"Foreign Banks",
    "CIMB BANK BERHAD":"Commercial Banks","CIMB INVESTMENT BANK BERHAD":"Investment Banks",
    "CIMB ISLAMIC BANK BERHAD":"Islamic Banks","CITIBANK BERHAD":"Foreign Banks",
    "DEUTSCHE BANK (MALAYSIA) BERHAD":"Foreign Banks",
    "DEUTSCHE BANK AG, INTERNATIONAL ISLAMIC BANKING, MALAYSIA BRANCH":"International Islamic Banks",
    "EXPORT-IMPORT BANK OF MALAYSIA BERHAD":"DFI","HONG LEONG BANK BERHAD":"Commercial Banks",
    "HONG LEONG INVESTMENT BANK BERHAD":"Investment Banks","HONG LEONG ISLAMIC BANK BERHAD":"Islamic Banks",
    "HSBC AMANAH MALAYSIA BERHAD":"Islamic Banks","HSBC BANK MALAYSIA BERHAD":"Commercial Banks",
    "IBDAR BANK BSC":"International Islamic Banks","INDIA INTERNATIONAL BANK (MALAYSIA) BERHAD":"Foreign Banks",
    "INDUSTRIAL AND COMMERCIAL BANK OF CHINA (MALAYSIA) BERHAD":"Foreign Banks",
    "J.P MORGAN CHASE BANK BERHAD":"Foreign Banks","KAF INVESTMENT BANK BERHAD":"Investment Banks",
    "KENANGA INVESTMENT BANK BERHAD":"Investment Banks","KUWAIT FINANCE HOUSE (MALAYSIA) BERHAD":"Islamic Banks",
    "MALAYAN BANKING BERHAD":"Commercial Banks","MAYBANK INVESTMENT BANK BERHAD":"Investment Banks",
    "MAYBANK ISLAMIC BERHAD":"Islamic Banks","MIDF AMANAH INVESTMENT BANK BERHAD":"Investment Banks",
    "MIZUHO CORPORATE BANK (MALAYSIA) BERHAD":"Foreign Banks",
    "NATIONAL BANK OF ABU DHABI MALAYSIA BERHAD":"Foreign Banks",
    "OCBC AL-AMIN BANK BERHAD":"Islamic Banks","OCBC BANK (MALAYSIA) BERHAD":"Commercial Banks",
    "PT. BANK SYARIAH MUALAMAT INDONESIA, TBK":"International Islamic Banks",
    "PUBLIC BANK BERHAD":"Commercial Banks","PUBLIC INVESTMENT BANK BERHAD":"Investment Banks",
    "PUBLIC ISLAMIC BANK BERHAD":"Islamic Banks","RHB BANK BERHAD":"Commercial Banks",
    "RHB INVESTMENT BANK BERHAD":"Investment Banks","RHB ISLAMIC BANK BERHAD":"Islamic Banks",
    "SMALL MEDIUM ENTERPRISE DEVELOPMENT BANK MALAYSIA BERHAD":"DFI",
    "STANDARD CHARTERED BANK MALAYSIA BERHAD":"Commercial Banks","STANDARD CHARTERED SAADIQ BERHAD":"Islamic Banks",
    "SUMITOMO MITSUI BANKING CORPORATION MALAYSIA BERHAD":"Foreign Banks",
    "THE BANK OF NOVA SCOTIA BERHAD":"Foreign Banks","THE ROYAL BANK OF SCOTLAND BERHAD":"Foreign Banks",
    "UNITED OVERSEAS BANK (MALAYSIA) BHD.":"Commercial Banks",
    "BOOST BANK BERHAD":"Digital Banks","AEON BANK (M) BERHAD":"Digital Banks","KAF DIGITAL BERHAD":"Digital Banks",
    "YTL DIGITAL BANK BERHAD":"Digital Banks","GX BANK BERHAD":"Digital Banks",
    # Insurers & Takaful (inc. reinsurers)
    "ACR RETAKAFUL BERHAD":"Takaful Operators","AIA BHD.":"Insurers","AIA GENERAL BERHAD":"Insurers",
    "AIA PUBLIC TAKAFUL BHD.":"Takaful Operators","AIG MALAYSIA INSURANCE BERHAD":"Insurers",
    "ALLIANZ GENERAL INSURANCE COMPANY (MALAYSIA) BERHAD":"Insurers",
    "ALLIANZ LIFE INSURANCE MALAYSIA BERHAD":"Insurers","AMBRA VERSICHERUNG AG":"Insurers",
    "AMMETLIFE INSURANCE BERHAD":"Insurers","AMMETLIFE TAKAFUL BERHAD":"Takaful Operators",
    "ASIA CAPITAL REINSURANCE MALAYSIA SDN. BHD.":"Insurers","BERJAYA SOMPO INSURANCE BERHAD":"Insurers",
    "CHUBB INSURANCE MALAYSIA BERHAD":"Insurers","DANAJAMIN NASIONAL BERHAD":"DFI",
    "ETIQA GENERAL INSURANCE BERHAD":"Insurers","ETIQA LIFE INSURANCE BERHAD":"Insurers",
    "ETIQA FAMILY TAKAFUL BERHAD":"Takaful Operators","ETIQA GENERAL TAKAFUL BERHAD":"Takaful Operators",
    "FWD INSURANCE BERHAD":"Insurers","FWD TAKAFUL BERHAD":"Takaful Operators",
    "GENERALI INSURANCE MALAYSIA BERHAD":"Insurers","GENERALI LIFE INSURANCE MALAYSIA BERHAD":"Insurers",
    "GREAT EASTERN GENERAL INSURANCE (MALAYSIA) BERHAD":"Insurers",
    "GREAT EASTERN LIFE ASSURANCE (MALAYSIA) BERHAD":"Insurers","GREAT EASTERN TAKAFUL BERHAD":"Takaful Operators",
    "HANNOVER RUECKVERSICHERUNG AG, MALAYSIAN BRANCH":"Insurers","HONG LEONG ASSURANCE BERHAD":"Insurers",
    "HONG LEONG MSIG TAKAFUL BERHAD":"Takaful Operators","LIBERTY GENERAL INSURANCE BERHAD":"Insurers",
    "LONPAC INSURANCE BHD.":"Insurers","MALAYSIAN LIFE REINSURANCE GROUP BERHAD":"Insurers",
    "MALAYSIAN REINSURANCE BERHAD":"Insurers","MANULIFE INSURANCE BERHAD":"Insurers",
    "MCIS INSURANCE BERHAD":"Insurers","MSIG INSURANCE (MALAYSIA) BHD":"Insurers",
    "MUNCHENER RUCKVERSICHERUNGS-GESELLSCHAFT (Munich Re Retakaful)":"Takaful Operators",
    "PACIFIC & ORIENT INSURANCE CO. BERHAD":"Insurers","PROGRESSIVE INSURANCE BHD.":"Insurers",
    "PRUDENTIAL ASSURANCE MALAYSIA BERHAD":"Insurers","PRUDENTIAL BSN TAKAFUL BERHAD":"Takaful Operators",
    "QBE INSURANCE (MALAYSIA) BERHAD":"Insurers","RHB INSURANCE BERHAD":"Insurers",
    "SUN LIFE MALAYSIA ASSURANCE BERHAD":"Insurers","SUN LIFE MALAYSIA TAKAFUL BERHAD":"Takaful Operators",
    "SWISS RE ASIA PTE LTD":"Insurers","SWISS REINSURANCE COMPANY LTD (SWISS RE RETAKAFUL)":"Takaful Operators",
    "SYARIKAT TAKAFUL MALAYSIA AM BERHAD":"Takaful Operators","SYARIKAT TAKAFUL MALAYSIA KELUARGA BERHAD":"Takaful Operators",
    "TAKAFUL IKHLAS FAMILY BERHAD":"Takaful Operators","TAKAFUL IKHLAS GENERAL BERHAD":"Takaful Operators",
    "THE PACIFIC INSURANCE BERHAD":"Insurers","THE TOA REINSURANCE COMPANY LTD.":"Insurers",
    "TOKIO MARINE INSURANS (MALAYSIA) BERHAD":"Insurers","TOKIO MARINE LIFE INSURANCE MALAYSIA BHD":"Insurers",
    "TUNE INSURANCE MALAYSIA BERHAD":"Insurers","ZURICH GENERAL INSURANCE MALAYSIA BERHAD":"Insurers",
    "ZURICH LIFE INSURANCE MALAYSIA BERHAD":"Insurers","ZURICH GENERAL TAKAFUL MALAYSIA BERHAD":"Takaful Operators",
    "ZURICH TAKAFUL MALAYSIA BERHAD":"Takaful Operators",
}

def attach_entity_type(df: pd.DataFrame, entity_col: str = "entity_name") -> pd.DataFrame:
    out = df.copy()
    out["_ENT_UP"] = out[entity_col].astype(str).str.upper().str.strip()
    out["entity_type"] = out["_ENT_UP"].map(ENTITY_TO_TYPE).fillna("Unknown")
    out.drop(columns=["_ENT_UP"], inplace=True)
    return out

def expand_rollups(
    df: pd.DataFrame,
    keys: Sequence[str],
    numeric_cols: Sequence[str],
    entity_col: str = "entity_name",
    entity_type_col: str = "entity_type",
    entity_group_col: str = "Entity / Group",
) -> pd.DataFrame:
    """
    Returns df with:
      - One row per unique (Entity/Group × other keys)
      - Rollup rows added once per key-combination
      - Only one display column 'Entity / Group' (no extra entity_type column)
    """
    if df.empty:
        return df

    # Base rows (FIs)
    base = df.copy()
    base[entity_group_col] = base[entity_col]

    out_frames = [base]

    # Rollup rows (aggregate by keys only, then assign the rollup label once)
    for roll_label, members in ROLLUPS.items():
        sub = base[base[entity_type_col].isin(members)]
        if sub.empty:
            continue
        agg = (
            sub.groupby(list(keys), dropna=False)[list(numeric_cols)]
               .sum(min_count=1).reset_index()
        )
        agg[entity_group_col] = roll_label
        out_frames.append(agg)

    out = pd.concat(out_frames, ignore_index=True)

    # Ensure single row per combo by summation (guard against dupes)
    gcols = [entity_group_col] + list(keys)
    out = out.groupby(gcols, dropna=False)[list(numeric_cols)].sum(min_count=1).reset_index()

    # Sort rollups first in desired order, then FIs A→Z
    rank = {n:i for i,n in enumerate(ROLLUP_ORDER)}
    out["_r"] = out[entity_group_col].map(lambda x: rank.get(x, 10_000))
    out = out.sort_values(["_r", entity_group_col] + list(keys), kind="mergesort").drop(columns=["_r"]).reset_index(drop=True)
    return out

# ========= Styling =========
FILL_HEADER = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_INFO   = PatternFill(start_color="EEF5FF", end_color="EEF5FF", fill_type="solid")
FILL_Q1     = PatternFill(start_color="E8F2FF", end_color="E8F2FF", fill_type="solid")
FILL_Q2     = PatternFill(start_color="EAFBEA", end_color="EAFBEA", fill_type="solid")
FILL_Q3     = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
FILL_Q4     = PatternFill(start_color="FCEEFF", end_color="FCEEFF", fill_type="solid")
FILL_RED    = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
FILL_YEL    = PatternFill(start_color="FFF8DB", end_color="FFF8DB", fill_type="solid")
BTHIN       = Border(left=Side(style="thin", color="DDDDDD"),
                     right=Side(style="thin", color="DDDDDD"),
                     top=Side(style="thin", color="DDDDDD"),
                     bottom=Side(style="thin", color="DDDDDD"))

def _q_fill(m: str):
    if m in Q_TO_MONTHS["Q1"]: return FILL_Q1
    if m in Q_TO_MONTHS["Q2"]: return FILL_Q2
    if m in Q_TO_MONTHS["Q3"]: return FILL_Q3
    if m in Q_TO_MONTHS["Q4"]: return FILL_Q4
    return FILL_HEADER

# ========= Workbook helpers =========
def new_empty_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def append_about_sheet_last(wb: Workbook, title: str, year: int, qlabel: str, notes: str = "") -> None:
    ws = wb.create_sheet("_About")
    ws["A1"] = title; ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Year"; ws["B3"] = int(year)
    ws["A4"] = "Quarter"; ws["B4"] = QMAP.get(str(qlabel).strip(), str(qlabel).strip())
    ws["A5"] = "Generated"; ws["B5"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    if notes:
        ws["A7"] = "Notes"; ws["B7"] = notes

# ========= Prior helpers (INDEX/MATCH; no XLOOKUP) =========
def _add_prior_sheet_quarters(
    wb: Workbook,
    prior_df: pd.DataFrame,
    key_cols: Sequence[str],
    qcols: Sequence[str],
    sheet_name: str
) -> Tuple[str, Dict[str, str]]:
    if prior_df is None or prior_df.empty:
        return (sheet_name, {})
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = "KEY"
    colmap: Dict[str,str] = {}
    usable = [q for q in ["Q1","Q2","Q3","Q4"] if q in qcols and q in prior_df.columns]
    for j, q in enumerate(usable, start=2):
        ws.cell(1, j, q)
        colmap[q] = ws.cell(1, j).column_letter
    for i, row in prior_df.iterrows():
        key = "|".join(str(row.get(k, "")).strip() for k in key_cols)
        ws.cell(i+2, 1, key)
        for j, q in enumerate(usable, start=2):
            ws.cell(i+2, j, row.get(q, 0) or 0)
    ws.sheet_state = "hidden"
    return (sheet_name, colmap)

def _add_prior_sheet_months(
    wb: Workbook,
    prior_df: pd.DataFrame,
    key_cols: Sequence[str],
    months: Sequence[str],
    sheet_name: str
) -> Tuple[str, Dict[str, str]]:
    if prior_df is None or prior_df.empty:
        return (sheet_name, {})
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = "KEY"
    colmap: Dict[str,str] = {}
    for j, m in enumerate(months, start=2):
        ws.cell(1, j, m)
        colmap[m] = ws.cell(1, j).column_letter
    for i, row in prior_df.iterrows():
        key = "|".join(str(row.get(k, "")).strip() for k in key_cols)
        ws.cell(i+2, 1, key)
        for j, m in enumerate(months, start=2):
            ws.cell(i+2, j, row.get(m, 0) or 0)
    ws.sheet_state = "hidden"
    return (sheet_name, colmap)

def _ensure_hidden_key_col(ws, header_row: int, first_data_row: int, last_data_row: int,
                           col_index: Dict[str,int], key_cols: Sequence[str], title: str) -> int:
    if title in col_index:
        return col_index[title]
    idx = len(col_index) + 1
    col_index[title] = idx
    ws.cell(header_row, idx, title)
    # Build ="A{r}&"|"&B{r}..."
    parts = []
    for name in key_cols:
        c = col_index[name]
        col_letter = ws.cell(1, c).column_letter
        parts.append(f'{col_letter}{{r}}')
    expr = "=" + '&"|"&'.join(parts)
    for r in range(first_data_row, last_data_row + 1):
        ws.cell(r, idx).value = expr.replace("{r}", str(r))
    ws.column_dimensions[ws.cell(1, idx).column_letter].hidden = True
    return idx

# ========= QC writer =========
@dataclass
class QCArgs:
    sheet_name: str
    df: pd.DataFrame                # staging slice for one question sheet
    wb: Workbook
    year: int
    current_q: str
    prior_df: Optional[pd.DataFrame]
    months_to_show: List[str]       # e.g., ["Jan","Feb","Mar"] or ["Apr","May","Jun"]...
    include_job_function: bool      # True only for Q4 job-functions (value column)
    yoy_quarters: Optional[List[str]] = None  # e.g., ["Q4"] for jobfunc
    mom_pct_threshold: float = 0.25
    qoq_pct_threshold: float = 0.25
    yoy_pct_threshold: float = 0.25
    abs_cutoff: float = 50.0

def write_qc_sheet(args: QCArgs) -> None:
    df = args.df.copy()
    if df.empty:
        return

    # Ensure optional cols exist
    for c in ["subquestion","worker_category","job_function"]:
        if c not in df.columns:
            df[c] = ""

    # Attach entity_type (internal only; final output shows a single Entity / Group column)
    df = attach_entity_type(df, "entity_name")

    # Select months to display, ensure numeric fields exist
    months = [m for m in args.months_to_show if m in MONTHS_FULL]
    for m in months:
        if m not in df.columns:
            df[m] = 0.0

    # Jobfunc mode → single quarter (Q4) from 'value'
    jobfunc_mode = bool(args.include_job_function)
    if jobfunc_mode and "value" not in df.columns:
        df["value"] = 0.0

    # Build rollups + collapse to single row per combo
    key_cols = ["subquestion","worker_category"]
    if jobfunc_mode:
        key_cols.append("job_function")

    numeric_cols = (["value"] if jobfunc_mode else months)
    df_roll = expand_rollups(df, keys=key_cols, numeric_cols=numeric_cols)

    # Prepare sheet
    ws = args.wb.create_sheet(args.sheet_name)

    # Info band
    ws["A1"] = args.sheet_name; ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "Year"; ws["B2"] = int(args.year)
    ws["A3"] = "Quarter"; ws["B3"] = QMAP.get(str(args.current_q).strip(), str(args.current_q).strip())
    ws["D2"] = "Threshold (±%)"; ws["E2"] = float(args.mom_pct_threshold)
    ws["D3"] = "Abs cutoff";     ws["E3"] = float(args.abs_cutoff)
    for rng in ("A1:A3","B2:B3","D2:E3"):
        for row in ws[rng]:
            for c in row:
                c.fill = FILL_INFO
                c.border = BTHIN

    # Column layout
    dims = ["Entity / Group","Subquestion","Worker Category"]
    if jobfunc_mode:
        dims.append("Job Function")

    headers: List[str] = []
    headers += dims

    # Data columns
    if jobfunc_mode:
        headers += ["Q4"]
    else:
        headers += months

        # Monthly Diff + MoM%
        month_diff_headers: List[str] = []
        month_pct_headers: List[str] = []
        for i in range(1, len(months)):
            m = months[i]
            month_diff_headers.append(f"Diff {m}")
            month_pct_headers.append(f"MoM {m}")
        headers += month_diff_headers + month_pct_headers

        # Quarter totals + QoQ
        q_sums: List[str] = []
        for q in ["Q1","Q2","Q3","Q4"]:
            need = set(Q_TO_MONTHS[q])
            if need.issubset(set(months)):
                q_sums.append(q)
        headers += q_sums
        for i in range(1, len(q_sums)):
            headers += [f"Diff {q_sums[i]}", f"%Diff {q_sums[i]}"]

    # YoY / Prior headers
    yoy_quarters = []
    yoy_months = []
    if jobfunc_mode:
        yoy_quarters = args.yoy_quarters or ["Q4"]
        headers += [f"Prior {q}" for q in yoy_quarters] + [f"YoY {q}" for q in yoy_quarters]
    else:
        if len(months) <= 2:
            yoy_months = months[:]
            headers += [f"Prior {m}" for m in months] + [f"YoY {m}" for m in months]
        else:
            q_sums_present = [q for q in ["Q1","Q2","Q3","Q4"] if set(Q_TO_MONTHS[q]).issubset(set(months))]
            yoy_quarters = q_sums_present
            headers += [f"Prior {q}" for q in yoy_quarters] + [f"YoY {q}" for q in yoy_quarters]

    # Header rows: quarter band row (above months), then real header row
    info_rows = 4
    band_row = info_rows + 1
    header_row = info_rows + 2

    # Write band row (empty row reserved for quarter labels)
    ws.append([])

    # Write the actual header row (ensures Excel recognizes headers → no “Column1”)
    ws.append(headers)
    col_index: Dict[str, int] = {h: i+1 for i, h in enumerate(headers)}

    # Style the header row cells
    for j, h in enumerate(headers, start=1):
        c = ws.cell(header_row, j)
        c.value = h
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = FILL_HEADER
        c.border = BTHIN

    # Quarter band labels above month columns (no merges → no flicker)
    if not jobfunc_mode:
        for m in months:
            j = col_index[m]
            ws.cell(band_row, j).value = next(q for q, ml in Q_TO_MONTHS.items() if m in ml)
            ws.cell(band_row, j).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(band_row, j).fill = _q_fill(m)
            ws.cell(band_row, j).border = BTHIN

    # Write data rows
    first_data_row = header_row + 1
    for _, r in df_roll.iterrows():
        row = [
            r.get("Entity / Group",""),
            r.get("subquestion",""),
            r.get("worker_category",""),
        ]
        if jobfunc_mode:
            row.append(r.get("job_function",""))

        if jobfunc_mode:
            row.append(r.get("value", 0))
        else:
            for m in months:
                row.append(r.get(m, 0))

            # placeholders for Diff + MoM%
            for _ in range(1, len(months)):
                row.append("")
            for _ in range(1, len(months)):
                row.append("")

            # placeholders for quarter sums & QoQ
            q_sums_present = [q for q in ["Q1","Q2","Q3","Q4"] if set(Q_TO_MONTHS[q]).issubset(set(months))]
            for _ in q_sums_present:
                row.append("")
            for _ in range(1, len(q_sums_present)):
                row.append("")  # Diff Qx
                row.append("")  # %Diff Qx

        # YoY placeholders
        if jobfunc_mode:
            for _ in yoy_quarters:
                row.append("")  # Prior Qx
            for _ in yoy_quarters:
                row.append("")  # YoY Qx
        else:
            if yoy_months:
                for _ in yoy_months:
                    row.append("")  # Prior m
                for _ in yoy_months:
                    row.append("")  # YoY m
            else:
                for _ in yoy_quarters:
                    row.append("")  # Prior Qx
                for _ in yoy_quarters:
                    row.append("")  # YoY Qx

        ws.append(row)

    last_data_row = first_data_row + len(df_roll) - 1
    if last_data_row < first_data_row:
        return

    # Freeze panes (after dims)
    freeze_col = 4 if not jobfunc_mode else 5
    ws.freeze_panes = ws.cell(first_data_row, freeze_col)

    # Apply table
    table_ref = f"A{header_row}:{ws.cell(last_data_row, len(headers)).coordinate}"
    tbl = Table(displayName=args.sheet_name.replace(" ","_"), ref=table_ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    ws.add_table(tbl)

    # Month header quarter tint & number formats
    if not jobfunc_mode:
        for m in months:
            j = col_index[m]
            ws.cell(header_row, j).fill = _q_fill(m)

    def _fmt_col(title: str, fmt: str):
        if title not in col_index: return
        j = col_index[title]
        for r in range(first_data_row, last_data_row + 1):
            ws.cell(r, j).number_format = fmt

    if jobfunc_mode:
        _fmt_col("Q4", "#,##0")
    else:
        for m in months:
            _fmt_col(m, "#,##0")
        for i in range(1, len(months)):
            _fmt_col(f"Diff {months[i]}", "#,##0")
            _fmt_col(f"MoM {months[i]}", "0.0%")

        q_sums_present = [q for q in ["Q1","Q2","Q3","Q4"] if set(Q_TO_MONTHS[q]).issubset(set(months))]
        for q in q_sums_present:
            _fmt_col(q, "#,##0")
        for i in range(1, len(q_sums_present)):
            _fmt_col(f"Diff {q_sums_present[i]}", "#,##0")
            _fmt_col(f"%Diff {q_sums_present[i]}", "0.0%")

    # ========= Formulas =========
    if not jobfunc_mode and len(months) >= 2:
        for i in range(1, len(months)):
            m = months[i]; prev = months[i-1]
            cd = col_index[f"Diff {m}"]
            cp = col_index[f"MoM {m}"]
            cm = col_index[m]
            pm = col_index[prev]
            for r in range(first_data_row, last_data_row + 1):
                a = ws.cell(r, cm).coordinate
                b = ws.cell(r, pm).coordinate
                ws.cell(r, cd).value = f"={a}-{b}"
                ws.cell(r, cp).value = f"=IF(AND({a}=0,{b}=0),0,IF({b}=0,\"N/A\",({a}-{b})/{b}))"

# --- quarter totals ---
if not jobfunc_mode:
    q_sums_present = [q for q in ["Q1","Q2","Q3","Q4"] if set(Q_TO_MONTHS[q]).issubset(set(months))]

    for r in range(first_data_row, last_data_row + 1):
        # read Subquestion cell text once (for Q4 check)
        sub_val = str(ws.cell(r, col_index.get("Subquestion", 0)).value or "").strip()

        # === Q1 ===
        if "Q1" in q_sums_present:
            if ws.title == "QC_Q1A_Main" or (
                ws.title == "QC_Q4"
                and "A. Number of Job Vacancies as at End of the Month" in sub_val
            ):
                # use end-month value (Mar)
                ref = ws.cell(r, col_index["Mar"]).coordinate
                ws.cell(r, col_index["Q1"]).value = f"={ref}"
            else:
                a = ws.cell(r, col_index["Jan"]).coordinate
                b = ws.cell(r, col_index["Mar"]).coordinate
                ws.cell(r, col_index["Q1"]).value = f"=SUM({a}:{b})"

        # === Q2 ===
        if "Q2" in q_sums_present:
            if ws.title == "QC_Q1A_Main" or (
                ws.title == "QC_Q4"
                and "A. Number of Job Vacancies as at End of the Month" in sub_val
            ):
                ref = ws.cell(r, col_index["Jun"]).coordinate
                ws.cell(r, col_index["Q2"]).value = f"={ref}"
            else:
                a = ws.cell(r, col_index["Apr"]).coordinate
                b = ws.cell(r, col_index["Jun"]).coordinate
                ws.cell(r, col_index["Q2"]).value = f"=SUM({a}:{b})"

        # === Q3 ===
        if "Q3" in q_sums_present:
            if ws.title == "QC_Q1A_Main" or (
                ws.title == "QC_Q4"
                and "A. Number of Job Vacancies as at End of the Month" in sub_val
            ):
                ref = ws.cell(r, col_index["Sep"]).coordinate
                ws.cell(r, col_index["Q3"]).value = f"={ref}"
            else:
                a = ws.cell(r, col_index["Jul"]).coordinate
                b = ws.cell(r, col_index["Sep"]).coordinate
                ws.cell(r, col_index["Q3"]).value = f"=SUM({a}:{b})"

        # === Q4 ===
        if "Q4" in q_sums_present:
            if ws.title == "QC_Q1A_Main" or (
                ws.title == "QC_Q4"
                and "A. Number of Job Vacancies as at End of the Month" in sub_val
            ):
                ref = ws.cell(r, col_index["Dec"]).coordinate
                ws.cell(r, col_index["Q4"]).value = f"={ref}"
            else:
                a = ws.cell(r, col_index["Oct"]).coordinate
                b = ws.cell(r, col_index["Dec"]).coordinate
                ws.cell(r, col_index["Q4"]).value = f"=SUM({a}:{b})"



        for i in range(1, len(q_sums_present)):
            q = q_sums_present[i]; pq = q_sums_present[i-1]
            cd = col_index[f"Diff {q}"]
            cp = col_index[f"%Diff {q}"]
            cq = col_index[q]
            pp = col_index[pq]
            for r in range(first_data_row, last_data_row + 1):
                a = ws.cell(r, cq).coordinate
                b = ws.cell(r, pp).coordinate
                ws.cell(r, cd).value = f"={a}-{b}"
                ws.cell(r, cp).value = f"=IF(AND({a}=0,{b}=0),0,IF({b}=0,\"N/A\",({a}-{b})/{b}))"

    # YoY using hidden prior sheet + hidden KEY
    # --- Build prior qc frame with the same keys/columns (robust for Q3: no subquestion) ---
    prior_qc_df = None
    if args.prior_df is not None and not args.prior_df.empty:
        p = args.prior_df.copy()

        # Ensure optional key columns exist in PRIOR too (Q3 has no subquestion)
        for c in ["subquestion", "worker_category", "job_function"]:
            if c not in p.columns:
                p[c] = ""

        # Ensure numeric columns expected for join exist in PRIOR as well
        if jobfunc_mode:
            if "value" not in p.columns:
                p["value"] = 0.0
        else:
            for m in months:
                if m not in p.columns:
                    p[m] = 0.0

        # Attach entity type and build rollups on PRIOR just like current
        p = attach_entity_type(p, "entity_name")
        p_keys = ["subquestion", "worker_category"]
        if jobfunc_mode:
            p_keys.append("job_function")
        p_nums = (["value"] if jobfunc_mode else months)
        p_roll = expand_rollups(p, keys=p_keys, numeric_cols=p_nums)

        # Map to display keys for prior sheet(s)
        p_roll["Entity / Group"] = p_roll["Entity / Group"]  # already created by expand_rollups
        p_roll["Subquestion"]     = p_roll["subquestion"]
        p_roll["Worker Category"] = p_roll["worker_category"]

        if jobfunc_mode:
            p_roll["Job Function"] = p_roll["job_function"]
            p_roll["Q4"] = p_roll["value"]
            prior_qc_df = p_roll[["Entity / Group","Subquestion","Worker Category","Job Function","Q4"]].copy()
        else:
            # Provide quarterly totals in prior if we’ll do YoY on quarters
            for q in ["Q1","Q2","Q3","Q4"]:
                need = set(Q_TO_MONTHS[q])
                if need.issubset(set(p_roll.columns)):
                    p_roll[q] = p_roll[list(need)].sum(axis=1)

            base_cols = ["Entity / Group","Subquestion","Worker Category"]
            keep_cols = base_cols + [c for c in ["Q1","Q2","Q3","Q4"] if c in p_roll.columns] \
                                   + [m for m in months if m in p_roll.columns]
            prior_qc_df = p_roll[keep_cols].copy()

    key_cols_display = ["Entity / Group","Subquestion","Worker Category"]
    if jobfunc_mode:
        key_cols_display.append("Job Function")

    if jobfunc_mode or (not jobfunc_mode and not months or len(months) > 2):
        yoy_qs = args.yoy_quarters or ([q for q in ["Q1","Q2","Q3","Q4"] if q in [c for c in headers if c in ["Q1","Q2","Q3","Q4"]]])
        psname = f"_{args.sheet_name}_PRIOR_Q"
        sheet, colmap = _add_prior_sheet_quarters(
            args.wb, prior_qc_df, key_cols_display, yoy_qs, psname
        )
        if colmap:
            key_idx = _ensure_hidden_key_col(ws, header_row, first_data_row, last_data_row,
                                             col_index, key_cols_display, "_KEY_Q")
            for q in yoy_qs:
                prior_hdr = f"Prior {q}"
                yoy_hdr   = f"YoY {q}"
                if prior_hdr not in col_index and yoy_hdr not in col_index:
                    continue
                if q not in colmap:
                    continue
                prior_col_letter = colmap[q]
                prior_range = f"{psname}!${prior_col_letter}:${prior_col_letter}"
                key_range   = f"{psname}!$A:$A"
                prior_out = col_index.get(prior_hdr)
                yoy_out   = col_index.get(yoy_hdr)
                curr_col  = col_index.get(q if not jobfunc_mode else "Q4")
                for r in range(first_data_row, last_data_row + 1):
                    key_addr = ws.cell(r, key_idx).coordinate
                    prior_expr = f"INDEX({prior_range},MATCH({key_addr},{key_range},0))"
                    if prior_out:
                        ws.cell(r, prior_out).value = f"=IFERROR({prior_expr},\"\")"
                        ws.cell(r, prior_out).number_format = "#,##0"
                    if yoy_out and curr_col:
                        curr_addr = ws.cell(r, curr_col).coordinate
                        ws.cell(r, yoy_out).value = (
                            f'=IFERROR(IF({prior_expr}="","N/A",'
                            f'IF(AND({prior_expr}=0,{curr_addr}=0),0,'
                            f'IF(AND({prior_expr}=0,{curr_addr}>0),"N/A",'
                            f'({curr_addr}-{prior_expr})/{prior_expr}))),"N/A")'
                        )
                        ws.cell(r, yoy_out).number_format = "0.0%"

    if (not jobfunc_mode) and months and len(months) <= 2:
        psname = f"_{args.sheet_name}_PRIOR_M"
        sheet, colmap = _add_prior_sheet_months(
            args.wb, prior_qc_df, key_cols_display, months, psname
        )
        if colmap:
            key_idx = _ensure_hidden_key_col(ws, header_row, first_data_row, last_data_row,
                                             col_index, key_cols_display, "_KEY_M")
            for m in months:
                prior_hdr = f"Prior {m}"
                yoy_hdr   = f"YoY {m}"
                if prior_hdr not in col_index and yoy_hdr not in col_index:
                    continue
                if m not in colmap:
                    continue
                prior_col_letter = colmap[m]
                prior_range = f"{psname}!${prior_col_letter}:${prior_col_letter}"
                key_range   = f"{psname}!$A:$A"
                prior_out = col_index.get(prior_hdr)
                yoy_out   = col_index.get(yoy_hdr)
                curr_col  = col_index.get(m)
                for r in range(first_data_row, last_data_row + 1):
                    key_addr = ws.cell(r, key_idx).coordinate
                    prior_expr = f"INDEX({prior_range},MATCH({key_addr},{key_range},0))"
                    if prior_out:
                        ws.cell(r, prior_out).value = f"=IFERROR({prior_expr},\"\")"
                        ws.cell(r, prior_out).number_format = "#,##0"
                    if yoy_out and curr_col:
                        curr_addr = ws.cell(r, curr_col).coordinate
                        ws.cell(r, yoy_out).value = (
                            f'=IFERROR(IF({prior_expr}="","N/A",'
                            f'IF(AND({prior_expr}=0,{curr_addr}=0),0,'
                            f'IF(AND({prior_expr}=0,{curr_addr}>0),"N/A",'
                            f'({curr_addr}-{prior_expr})/{prior_expr}))),"N/A")'
                        )
                        ws.cell(r, yoy_out).number_format = "0.0%"

    # ========= Conditional formatting =========
    def _cf_percent(col_title: str, diff_title: Optional[str], pct_threshold: float, abs_cutoff: float):
        if col_title not in col_index:
            return
        j = col_index[col_title]
        col_letter = ws.cell(header_row, j).column_letter
        rng = f"{col_letter}{first_data_row}:{col_letter}{last_data_row}"
        top = f"{col_letter}{first_data_row}"
        if diff_title and diff_title in col_index:
            dj = col_index[diff_title]
            dletter = ws.cell(header_row, dj).column_letter
            dtop = f"{dletter}{first_data_row}"
            red = f"=AND(ISNUMBER({top}),ABS({top})>={pct_threshold},ABS({dtop})>={abs_cutoff})"
            yel = f"=AND(ISNUMBER({top}),ABS({top})>={pct_threshold},ABS({dtop})<{abs_cutoff})"
        else:
            red = f"=AND(ISNUMBER({top}),ABS({top})>={pct_threshold},ABS({top})>={pct_threshold})"
            yel = f"=AND(ISNUMBER({top}),ABS({top})>={pct_threshold},ABS({top})<{pct_threshold})"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[red], fill=FILL_RED))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[yel], fill=FILL_YEL))

    if not jobfunc_mode and len(months) >= 2:
        for i in range(1, len(months)):
            m = months[i]
            _cf_percent(f"MoM {m}", f"Diff {m}", args.mom_pct_threshold, args.abs_cutoff)

        q_sums_present = [q for q in ["Q1","Q2","Q3","Q4"] if set(Q_TO_MONTHS[q]).issubset(set(months))]
        for i in range(1, len(q_sums_present)):
            q = q_sums_present[i]
            _cf_percent(f"%Diff {q}", f"Diff {q}", args.qoq_pct_threshold, args.abs_cutoff)

    if jobfunc_mode:
        for q in yoy_quarters:
            _cf_percent(f"YoY {q}", None, args.yoy_pct_threshold, args.abs_cutoff)
    else:
        if yoy_months:
            for m in yoy_months:
                _cf_percent(f"YoY {m}", None, args.yoy_pct_threshold, args.abs_cutoff)
        else:
            for q in yoy_quarters:
                _cf_percent(f"YoY {q}", None, args.yoy_pct_threshold, args.abs_cutoff)

    # ========= Auto-width (light) =========
    for col in range(1, len(headers) + 1):
        maxw = max(len(str(ws.cell(header_row, col).value or "")), 10)
        for r in range(first_data_row, min(first_data_row + 50, last_data_row + 1)):
            v = ws.cell(r, col).value
            if v is not None:
                maxw = max(maxw, len(str(v)))
        ws.column_dimensions[ws.cell(1, col).column_letter].width = min(maxw + 2, 28)
