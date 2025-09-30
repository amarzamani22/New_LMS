from __future__ import annotations
import argparse, time
from pathlib import Path
from typing import List, Dict, Tuple, Optional
import pandas as pd
from openpyxl import load_workbook

# ---- Fixed layout for Q3 (rows & columns) ----
# Worker categories appear in these exact rows for Q3
Q3_ROWS = [
    (169, "Managers"),
    (170, "Professional"),
    (171, "Technicians & Associate Professionals"),
    (172, "Clerical Occupations"),
    (173, "Operative Workers"),
    (174, "Elementary Occupations"),
    (175, "TOTAL Hours Worked During the Month, Including Overtime"),
]

# Values are always in C, D, E for the active quarter
VAL_COLS = ["C", "D", "E"]

# Preferred data sheet name; we’ll fall back to first non-Cover if missing
PREFERRED_SHEET = "Banking & DFI"

# Map the various quarter labels on Cover!F8 to month names
Q_TO_MONTHS: Dict[str, Tuple[str, str, str]] = {
    "Q1": ("Jan","Feb","Mar"),
    "QUARTER 1": ("Jan","Feb","Mar"),
    "Q2": ("Apr","May","Jun"),
    "QUARTER 2": ("Apr","May","Jun"),
    "Q3": ("Jul","Aug","Sep"),
    "QUARTER 3": ("Jul","Aug","Sep"),
    "Q4": ("Oct","Nov","Dec"),
    "QUARTER 4": ("Oct","Nov","Dec"),
}

def read_cover(wb) -> Tuple[Optional[str], Optional[int], Optional[str]]:
    """Return (entity, year, quarter_label) from Cover sheet cells F6/F7/F8."""
    ent = yr = q = None
    if "Cover" in wb.sheetnames:
        ws = wb["Cover"]
        ent = ws["F6"].value
        yr  = ws["F7"].value
        q   = ws["F8"].value
    # normalize
    ent = str(ent).strip() if ent not in (None, "") else None
    try:
        yr = int(str(yr).strip()) if yr not in (None, "") else None
    except Exception:
        yr = None
    q = str(q).strip() if q not in (None, "") else None
    return ent, yr, q

def pick_data_sheet(wb) -> str:
    if PREFERRED_SHEET in wb.sheetnames:
        return PREFERRED_SHEET
    for s in wb.sheetnames:
        if s != "Cover":
            return s
    return wb.sheetnames[0]

def read_num(ws, addr: str) -> float:
    v = ws[addr].value
    if v in (None, "", "-"):
        return 0.0
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).replace(",", ""))
        except Exception:
            return 0.0

def extract_q3_from_file(path: Path) -> pd.DataFrame:
    try:
        wb = load_workbook(str(path), data_only=True, read_only=True)
    except Exception:
        return pd.DataFrame()

    ent, yr, q = read_cover(wb)
    if not (ent and yr and q):
        try: wb.close()
        except Exception: pass
        return pd.DataFrame()

    months = Q_TO_MONTHS.get(q.upper())
    if not months:  # unrecognized quarter label
        try: wb.close()
        except Exception: pass
        return pd.DataFrame()

    ws = wb[pick_data_sheet(wb)]

    rows: List[Dict] = []
    m1, m2, m3 = months
    for r, wc in Q3_ROWS:
        v1 = read_num(ws, f"{VAL_COLS[0]}{r}")
        v2 = read_num(ws, f"{VAL_COLS[1]}{r}")
        v3 = read_num(ws, f"{VAL_COLS[2]}{r}")
        rows.append({
            "entity_name": ent,
            "year": yr,
            "quarter": q,
            "question": "Q3",
            "worker_category": wc,
            m1: v1,
            m2: v2,
            m3: v3,
        })

    try: wb.close()
    except Exception: pass

    return pd.DataFrame(rows)

def main() -> int:
    ap = argparse.ArgumentParser(description="Extract Question 3 (fixed cells, fast).")
    ap.add_argument("--input", required=True, help="Folder with submissions (.xlsx/.xlsm)")
    ap.add_argument("--out",   required=True, help="Output staging workbook (.xlsx)")
    ap.add_argument("--limit", type=int, default=None, help="Limit files (debug)")
    args = ap.parse_args()

    root = Path(args.input)
    files = []
    for ext in ("*.xlsx","*.xlsm"):
        files.extend(p for p in root.rglob(ext) if not p.name.startswith("~$"))
    files.sort()
    if args.limit:
        files = files[:args.limit]
    print(f"[INFO] Files: {len(files)}")

    t0 = time.perf_counter()
    frames: List[pd.DataFrame] = []
    for p in files:
        df = extract_q3_from_file(p)
        if not df.empty:
            frames.append(df)

    out_df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame(
        columns=["entity_name","year","quarter","question","worker_category","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"]
    )

    # Keep only month columns that actually appear
    base = ["entity_name","year","quarter","question","worker_category"]
    month_cols = [c for c in ["Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"] if c in out_df.columns]
    out_df = out_df[base + month_cols] if not out_df.empty else out_df

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(out_path, engine="openpyxl") as xw:
        out_df.to_excel(xw, index=False, sheet_name="Q3")

    print(f"[DONE] Wrote → {out_path} in {time.perf_counter()-t0:0.2f}s")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())