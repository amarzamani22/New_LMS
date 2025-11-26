def mhs_page():
    """
    FINAL VERSION (REAL ONE):
    - User selects YEAR + QUARTER
    - System extracts QC totals EXACTLY like main app:
        • match subquestion via .contains()
        • filter Entity == 'All Financial Institutions'
        • sum across ALL Worker Categories
    - Auto-fills 3 months + quarter row
    - Auto-fills Year row only for Q4
    - Shows monthly + quarter preview tables
    """
    import io
    import openpyxl
    from openpyxl.utils import get_column_letter
    import pandas as pd
    from datetime import datetime

    st.header("MHS Table Generator — Quarter Auto-Fill")

    # --------------------------------------------------
    # 1. USER INPUTS
    # --------------------------------------------------
    now_year = datetime.now().year
    years = list(range(now_year - 5, now_year + 3))
    year = st.selectbox("Select Year", years, index=years.index(now_year))
    quarter = st.selectbox("Select Quarter", ["Q1", "Q2", "Q3", "Q4"])

    quarter_map = {
        "Q1": [1, 2, 3],
        "Q2": [4, 5, 6],
        "Q3": [7, 8, 9],
        "Q4": [10, 11, 12]
    }

    q_months = quarter_map[quarter]

    template_file = st.file_uploader("Upload MHS Template (.xlsx)", type=["xlsx"])
    if not template_file:
        return

    # --------------------------------------------------
    # 2. LOAD TEMPLATE
    # --------------------------------------------------
    wb = openpyxl.load_workbook(template_file)
    ws = wb[wb.sheetnames[0]]

    # anchors (same as before)
    ANCHOR_MONTH = 165   # C165 start
    ANCHOR_QUARTER = 55  # C55 start
    ANCHOR_YEAR = 15     # B15 start

    START_COL = 4  # Column D is index 4

    # --------------------------------------------------
    # 3. UTILS
    # --------------------------------------------------
    def next_empty(sheet, row, col_letter):
        while sheet[f"{col_letter}{row}"].value not in (None, ""):
            row += 1
        return row

    def month_exists(sheet, y, m):
        r = ANCHOR_MONTH
        while r < 5000:
            cell = sheet[f"C{r}"].value
            if cell == m:
                # verify year above
                for back in range(0, 10):
                    yc = sheet[f"B{r-back}"].value
                    if yc == y:
                        return True
            r += 1
        return False

    # --------------------------------------------------
    # 4. LOAD QC SHEETS
    # --------------------------------------------------
    try:
        df_q1 = load_qc_sheet(year, SHEET_MAP["Q1A: Employees"])
        df_q4 = load_qc_sheet(year, SHEET_MAP["Q4: Vacancies"])
        df_q5 = load_qc_sheet(year, SHEET_MAP["Q5: Separations"])
    except Exception as e:
        st.error(f"Error loading QC: {e}")
        return

    # --------------------------------------------------
    # 5. REAL EXACT QC EXTRACTION (MATCHES YOUR APP)
    # --------------------------------------------------
    def qc(df, keyword, month_idx):
        """
        EXACT extract:
        - match Subquestion using .contains()
        - filter Entity == 'All Financial Institutions'
        - sum across ALL worker categories
        """
        if df is None or isinstance(df, str):
            return 0.0

        # Subquestion match (NOT exact match)
        sel = df[df["Subquestion"].astype(str).str.contains(keyword, case=False, na=False)]
        if sel.empty:
            return 0.0

        # Entity filtering EXACTLY like your app
        ent_col = None
        for c in df.columns:
            if c.lower().replace(" ", "") in ["entity/group", "entity/group", "entity/group"]:
                ent_col = c
                break
            if "entity" in c.lower():
                ent_col = c
                break

        if ent_col:
            sel = sel[sel[ent_col].astype(str).str.strip() == "All Financial Institutions"]

        if sel.empty:
            return 0.0

        # Sum across all worker categories
        month_col = ALL_MONTHS[month_idx - 1]
        try:
            return float(pd.to_numeric(sel[month_col], errors="coerce").fillna(0).sum())
        except:
            return 0.0

    # --------------------------------------------------
    # 6. SUBQUESTION KEYWORDS (SAFE VERSION)
    # --------------------------------------------------
    SUB = {
        "EMP": ["Number of Employees", "Employees"],
        "VAC": ["Vacancies"],
        "NEWJOB": ["New Jobs", "New Job"],
        "HIRE": ["Hire", "Hires"],
        "QUIT": ["Quit"],
        "LAYOFF": ["Layoff"],
        "OTHER": ["Other Separation", "Other"],
    }

    def get_val(df, keys, m):
        """Try multiple keywords until one matches."""
        for k in keys:
            v = qc(df, k, m)
            if v != 0:
                return v
        return 0.0

    # --------------------------------------------------
    # 7. BUILD MONTH ROWS
    # --------------------------------------------------
    month_rows = []
    for m in q_months:
        EMP = (
            get_val(df_q1, SUB["EMP"], m)
        )
        VAC = get_val(df_q4, SUB["VAC"], m)
        NEWJ = get_val(df_q4, SUB["NEWJOB"], m)
        HIRE = get_val(df_q5, SUB["HIRE"], m)
        QUIT = get_val(df_q5, SUB["QUIT"], m)
        LAY = get_val(df_q5, SUB["LAYOFF"], m)
        OTH = get_val(df_q5, SUB["OTHER"], m)
        SEP = QUIT + LAY + OTH

        month_rows.append([m, EMP, VAC, NEWJ, HIRE, SEP, QUIT, LAY, OTH])

    # --------------------------------------------------
    # 8. PREVIEW TABLES
    # --------------------------------------------------
    st.subheader("📅 Monthly Breakdown")
    df_prev = pd.DataFrame(month_rows, columns=[
        "Month", "Employment", "Vacancies", "New Jobs",
        "Hires", "Separations Total", "Quits", "Layoff", "Other"
    ])
    st.dataframe(df_prev, use_container_width=True)

    st.subheader("📊 Quarter Summary")
    quarter_totals = df_prev.iloc[:, 1:].sum()
    st.dataframe(quarter_totals.to_frame("Value"))

    # --------------------------------------------------
    # 9. WRITE TO EXCEL
    # --------------------------------------------------
    if st.button("Write Quarter to MHS Template"):
        # check if any month exists
        for m in q_months:
            if month_exists(ws, year, m):
                st.error(f"❌ Month {m} {year} already exist.")
                return

        # Write 3 months
        for m, *vals in month_rows:
            r = next_empty(ws, ANCHOR_MONTH, "C")
            ws[f"B{r}"] = year
            ws[f"C{r}"] = m
            for i, v in enumerate(vals):
                ws[f"{get_column_letter(START_COL + i)}{r}"] = v

        # Write quarter row
        qr = next_empty(ws, ANCHOR_QUARTER, "C")
        ws[f"B{qr}"] = year
        ws[f"C{qr}"] = quarter
        for i, v in enumerate(quarter_totals):
            ws[f"{get_column_letter(START_COL + i)}{qr}"] = float(v)

        # Write year row only for Q4
        if quarter == "Q4":
            yr = next_empty(ws, ANCHOR_YEAR, "B")
            year_total_vals = df_prev.iloc[:, 1:].sum()
            ws[f"B{yr}"] = year
            for i, v in enumerate(year_total_vals):
                ws[f"{get_column_letter(START_COL + i)}{yr}"] = float(v)

        # Save back
        out = io.BytesIO()
        wb.save(out)

        st.success("Quarter inserted successfully!")
        st.download_button(
            "Download Updated MHS File",
            data=out.getvalue(),
            file_name=f"MHS_{year}_{quarter}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
