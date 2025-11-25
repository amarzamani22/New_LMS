#----------------------------------------------------
# LMS Analysis Dashboard (Streamlit)
#----------------------------------------------------
# 1)This code loads QC workbooks for the LMS
# 2)Detects statistically significant outliers, and links them to Variance Report
# 3)Show list of FI that contribute to the detected outliers
#-----------------------------------------------------

import os
import re
from typing import Dict, List, Optional, Tuple
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
import streamlit as st
from openpyxl.utils import get_column_letter
from datetime import datetime
import openpyxl
import io

# ================ Page & Styles in CSS =================
st.set_page_config(page_title="LMS Analysis Dashboard", page_icon="🏦", layout="wide")
st.markdown("""""", unsafe_allow_html=True)

# ================ Constants / Configuration ================

# Mapping datasets name to specific QC sheet names in Excel workbook
SHEET_MAP: Dict[str, str] = {
    'Q1A: Employees': 'QC_Q1A_Main',
    'Q2A: Salary': 'QC_Q2A_Main',
    'Q3: Hours Worked': 'QC_Q3',
    'Q4: Vacancies': 'QC_Q4',
    'Q5: Separations': 'QC_Q5'
}
ALL_MONTHS: List[str] = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
ENTITY_COL = "Entity / Group"                 #Column name used to identify entities or groups in QC sheet
SUBQ_COL = "Subquestion"                      #Column name for subquestion (except for Question 3(Hours Worked))
WC_COL = "Worker Category"                    #Column name for worker category
ROLLUP_KEY = "All Financial Institutions"     #Roll-up label

# ---- Minimal fallback hierarchy (used only if no CSV is provided) ----
HIERARCHY_MAP: Dict[str, List[str]] = {
    "All Financial Institutions": ["Banking Institution", "DFI", "Insurans/Takaful Operators"],
    "Banking Institution": [
        "Commercial Banks", "Digital Banks", "Foreign Banks",
        "Islamic Banks", "Investment Banks", "International Islamic Banks"
    ],
    "DFI": [],  # Actual DFI list will come from CSV when available
    "Insurans/Takaful Operators": ["Insurers", "Takaful Operators"],
}

# ====================== Data Access (QC Workbook) ======================
@st.cache_data
def load_qc_sheet(year: int, sheet_name: str) -> pd.DataFrame | str:
    #Load a QC sheet for a give year and sheet name
    #Returns a DataFrame on success, or an error string on failure
    #Caches results to avoid re-reading repeatedly
    path = fr"C:\Users\ttamarz\OneDrive - Bank Negara Malaysia\RLMS\Output\QC Template\qc_workbook_{year}.xlsx"

    if not os.path.exists(path):
        return f"Error: File not found → {path}"
    try:
        # QC sheets have 5 header rows above the data header
        df = pd.read_excel(path, sheet_name=sheet_name, header=5)

        # Remove empty columns
        df.dropna(axis=1, how='all', inplace=True)

        # Normalize quarter total column names
        df.rename(columns={'Q1.1':'Q1_Total','Q2.1':'Q2_Total','Q3.1':'Q3_Total','Q4.1':'Q4_Total'}, inplace=True)
        return df
    except Exception as e:
        return f"Error reading sheet '{sheet_name}' from {path}: {e}"
    
# ============ Helpers for MHS =================
def mhs_page():
    """
    Streamlit page: MHS Table Generator (separate tab).
    Relies on existing functions/consts in your app:
      - load_qc_sheet(year, sheet_name)
      - SHEET_MAP  (mapping of question -> sheet name)
      - ALL_MONTHS (['Jan','Feb',...])
    Paste this function in App.py and call it from your nav when 'MHS Table' selected.
    """

    st.header("MHS Table Generator — Monthly Highlight Statistics")
    st.write("Select a month to fill into an MHS Excel template. The system will pull values directly from QC sheets (Q1A, Q4, Q5) for the selected year and write Month → Quarter → Year rows automatically.")

    # Allow uploading the MHS template (the same template you use)
    mhs_file = st.file_uploader("Upload MHS template (.xlsx) — the same template structure used by reports", type=["xlsx"])
    if not mhs_file:
        st.info("Upload your MHS template file (e.g. 3.5.12a.xlsx).")
        return

    # Optionally show example path to test (we keep sample path variable if you want to test quickly)
    SAMPLE_TEMPLATE_PATH = "/mnt/data/3.5.12a.xlsx"  # <--- local sample, optional

    # Target month input
    ym = st.text_input("Target month to add (YYYY-MM)", value=datetime.now().strftime("%Y-%m"))
    try:
        target_year = int(ym.split("-")[0])
        target_month = int(ym.split("-")[1])
        if not (1 <= target_month <= 12):
            raise ValueError
    except Exception:
        st.error("Target month must be in YYYY-MM format (e.g. 2025-03).")
        return

    # Fill preference
    fill_option = st.selectbox("Fill behaviour", [
        "Month only",
        "Month + Quarter if complete",
        "Month + Quarter + Year if complete"
    ])

    # Anchor cells as per your template
    ANCHOR = {
        "year_block_row": 15,        # B15 anchors the Year block label area
        "quarter_year_row": 53,      # B53 anchors the Quarter-year area label
        "quarter_first_row": 55,     # C55 first quarter row anchor
        "month_year_row": 165,       # B165 anchors Year in Month block
        "month_first_row": 165       # C165 first month cell
    }

    # Convenience: month name from month number (use your ALL_MONTHS constant)
    try:
        month_name = ALL_MONTHS[target_month - 1]
    except Exception:
        # fallback
        month_name = datetime(2000, target_month, 1).strftime("%b")

    st.write(f"Preparing to add **{month_name} {target_year}**")

    # Utility helpers --------------------------------------------------
    def next_empty_row(sheet, start_row:int, col_letter:str):
        r = start_row
        while sheet[f"{col_letter}{r}"].value not in (None, ""):
            r += 1
            if r > 5000:
                raise RuntimeError("No empty row found - template shape unexpected.")
        return r

    def month_exists_in_template(sheet, year:int, month:int) -> bool:
        # scan month column (C) in month block to see (month,year) pair exists
        r = ANCHOR["month_first_row"]
        while r < 5000:
            cell = sheet[f"C{r}"].value
            if cell is None or cell == "":
                r += 1
                continue
            # if cell equals numeric month label (1..12) or name
            try:
                # if month stored as number
                if int(cell) == month:
                    # find nearest year above (col B)
                    up = r
                    found_year = None
                    for back in range(0, 12):
                        ycell = sheet[f"B{up-back}"].value
                        if isinstance(ycell, int):
                            found_year = ycell
                            break
                        if isinstance(ycell, str) and ycell.strip().isdigit() and len(ycell.strip())==4:
                            found_year = int(ycell.strip()); break
                    if found_year == year:
                        return True
            except Exception:
                # maybe stored as month name (e.g., "Jan"), compare names
                if isinstance(cell, str) and cell.strip().lower().startswith(ALL_MONTHS[month-1][:3].lower()):
                    up = r
                    found_year = None
                    for back in range(0, 12):
                        ycell = sheet[f"B{up-back}"].value
                        if isinstance(ycell, int):
                            found_year = ycell; break
                        if isinstance(ycell, str) and ycell.strip().isdigit() and len(ycell.strip())==4:
                            found_year = int(ycell.strip()); break
                    if found_year == year:
                        return True
            r += 1
        return False

    # Load MHS template workbook via openpyxl (preserve formatting)
    try:
        wb = openpyxl.load_workbook(mhs_file)
        ws = wb[wb.sheetnames[0]]
    except Exception as e:
        st.error(f"Failed to read uploaded template: {e}")
        return

    # Guard: abort if month exists (Option C)
    if month_exists_in_template(ws, target_year, target_month):
        st.error(f"{ym} already exists in the MHS template. Operation aborted (per Option C).")
        return

    # Fetch QC sheets for target year using your existing loader -----------------
    # Uses your load_qc_sheet function and SHEET_MAP values defined in App.py
    try:
        q1a_sheetname = SHEET_MAP['Q1A: Employees']
        q4_sheetname  = SHEET_MAP['Q4: Vacancies']
        q5_sheetname  = SHEET_MAP['Q5: Separations']

        df_q1 = load_qc_sheet(target_year, q1a_sheetname)
        df_q4 = load_qc_sheet(target_year, q4_sheetname)
        df_q5 = load_qc_sheet(target_year, q5_sheetname)
    except Exception as e:
        st.error(f"Failed to load QC sheets for year {target_year}: {e}")
        return

    # Subquestion labels (exact strings extracted from your mapping)
    SUBQ_MAP = {
        "emp_A": "A. Number of Employees",
        "emp_B1": "B(i). Malaysian Employees",
        "emp_B2": "B(ii). Non-Malaysian Employees",

        "vac_total": "A. Number of Job Vacancies as at End of the Month",
        "vac_newjob": "Number of Job Vacancies Due to New Jobs Created During the Month",

        "hires_recalls": "New Hires and Recalls",

        "sep_quit": "A. Quits and resignation (except retirement)",
        "sep_layoff": "B. Total Layoffs and Discharges",
        "sep_other": "C. Other Separation"
    }

    # Helper to pick a numeric value from a QC DF for a subquestion and month
    def qc_month_value(df, subq_label, month_idx:int):
        # df expected to have 'Subquestion' column and Jan..Dec columns
        if isinstance(df, str):
            # load_qc_sheet may return error string
            return 0
        try:
            sel = df[df["Subquestion"].astype(str).str.strip() == subq_label]
            if sel.empty:
                # Sometimes subquestion labels differ slightly; try contains fallback
                sel = df[df["Subquestion"].astype(str).str.contains(subq_label.split()[0], na=False)]
            if sel.empty:
                return 0
            colname = ALL_MONTHS[month_idx-1]  # "Jan".."Dec"
            # sum across any matching rows and return numeric (tolerant)
            vals = pd.to_numeric(sel[colname], errors='coerce').fillna(0).sum()
            return float(vals)
        except Exception:
            return 0

    # Compute monthly indicator values (use qc_month_value)
    emp_A = qc_month_value(df_q1, SUBQ_MAP["emp_A"], target_month)
    emp_B1 = qc_month_value(df_q1, SUBQ_MAP["emp_B1"], target_month)
    emp_B2 = qc_month_value(df_q1, SUBQ_MAP["emp_B2"], target_month)
    total_employment = emp_A + emp_B1 + emp_B2

    vac_total = qc_month_value(df_q4, SUBQ_MAP["vac_total"], target_month)
    vac_newjob = qc_month_value(df_q4, SUBQ_MAP["vac_newjob"], target_month)

    hires = qc_month_value(df_q5, SUBQ_MAP["hires_recalls"], target_month)

    sep_quit = qc_month_value(df_q5, SUBQ_MAP["sep_quit"], target_month)
    sep_layoff = qc_month_value(df_q5, SUBQ_MAP["sep_layoff"], target_month)
    sep_other = qc_month_value(df_q5, SUBQ_MAP["sep_other"], target_month)
    sep_total = sep_quit + sep_layoff + sep_other

    # Final ordered list of indicator values to write horizontally.
    # NOTE: adjust order if your MHS template expects different column order.
    indicator_values = [
        total_employment,
        vac_total,
        vac_newjob,
        hires,
        sep_total,
        sep_quit,
        sep_layoff,
        sep_other
    ]

    st.write("Computed values (will be written left→right into the template indicator columns):")
    preview_df = pd.DataFrame({
        "indicator": [
            "Total Employment", "Vacancies (total)", "Vacancies (new job)", "New hires & recalls",
            "Separations (total)", "Separations (quit)", "Separations (layoff)", "Separations (other)"
        ],
        "value": [round(x, 2) for x in indicator_values]
    })
    st.dataframe(preview_df, use_container_width=True)

    # Where to write indicators in the MHS template?
    # We assume indicators start at column D (col index 4). If your template differs,
    # change start_col_idx accordingly.
    start_col_idx = 4  # Column D

    # Confirm and write
    if st.button("Write MHS row and generate file"):
        # write year label row if not present near month block
        # find first empty row in column B after month_year_row anchor
        try:
            yrow = next_empty_row(ws, ANCHOR["month_year_row"], "B")
            ws[f"B{yrow}"].value = target_year
        except Exception:
            # if B cell already contains year on top of month block, it's fine
            pass

        # write month entry
        mrow = next_empty_row(ws, ANCHOR["month_first_row"], "C")
        ws[f"C{mrow}"].value = target_month

        # write indicator values horizontally starting at start_col_idx
        for i, v in enumerate(indicator_values):
            colletter = get_column_letter(start_col_idx + i)
            ws[f"{colletter}{mrow}"].value = v

        # QUARTER: if requested and quarter complete -> sum quarter months and insert
        def present_months_for_year(sheet, year):
            months_present = set()
            r = ANCHOR["month_first_row"]
            while r < 5000:
                val = sheet[f"C{r}"].value
                if val in (None, ""):
                    r += 1; continue
                try:
                    mm = int(val)
                except Exception:
                    # maybe month name
                    try:
                        mm = ALL_MONTHS.index(str(val)[:3]) + 1
                    except Exception:
                        r += 1; continue
                # get year above
                found_year = None
                for back in range(0, 12):
                    ycell = sheet[f"B{r-back}"].value
                    if isinstance(ycell, int):
                        found_year = ycell; break
                    if isinstance(ycell, str) and ycell.strip().isdigit() and len(ycell.strip())==4:
                        found_year = int(ycell.strip()); break
                if found_year == year:
                    months_present.add(mm)
                r += 1
            return months_present

        if fill_option != "Month only":
            quarter_idx = (target_month - 1) // 3 + 1
            qmonths = {(quarter_idx - 1) * 3 + i for i in (1,2,3)}
            months_present = present_months_for_year(ws, target_year)
            if qmonths.issubset(months_present):
                # compute quarter sums by reading month rows we just wrote (or previously present)
                qvals = []
                for i in range(len(indicator_values)):
                    total = 0
                    # scan for each month in qmonths and add the value at column start_col_idx + i
                    for m in qmonths:
                        r = ANCHOR["month_first_row"]
                        while r < 5000:
                            cellm = ws[f"C{r}"].value
                            if cellm == m:
                                colletter = get_column_letter(start_col_idx + i)
                                total += ws[f"{colletter}{r}"].value or 0
                                break
                            r += 1
                    qvals.append(total)
                # insert quarter row at next empty row under quarter_first_row (col C)
                qrow = next_empty_row(ws, ANCHOR["quarter_first_row"], "C")
                ws[f"C{qrow}"].value = f"{quarter_idx}Q"
                ws[f"B{qrow}"].value = target_year
                for i, v in enumerate(qvals):
                    colletter = get_column_letter(start_col_idx + i)
                    ws[f"{colletter}{qrow}"].value = v

        # YEAR: if requested and all 12 months present -> sum and insert at year block
        if fill_option == "Month + Quarter + Year if complete":
            months_present = present_months_for_year(ws, target_year)
            if set(range(1,13)).issubset(months_present):
                yvals = []
                for i in range(len(indicator_values)):
                    total = 0
                    for m in range(1,13):
                        r = ANCHOR["month_first_row"]
                        while r < 5000:
                            if ws[f"C{r}"].value == m:
                                total += ws[f"{get_column_letter(start_col_idx + i)}{r}"].value or 0
                                break
                            r += 1
                    yvals.append(total)
                yrow = next_empty_row(ws, ANCHOR["year_block_row"], "B")
                ws[f"B{yrow}"].value = target_year
                for i, v in enumerate(yvals):
                    ws[f"{get_column_letter(start_col_idx + i)}{yrow}"].value = v

        # Save workbook to bytes and provide download
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        st.success("MHS template updated. Download ready.")
        st.download_button(
            label="Download updated MHS Excel",
            data=out.getvalue(),
            file_name=f"MHS_updated_{target_year}_{target_month}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

@st.cache_data
def get_reporting_quarter(year: int) -> int:

    #Read _About sheet to know the current quarter in that workbook.
    path = fr"C:\Users\ttamarz\OneDrive - Bank Negara Malaysia\RLMS\Output\QC Template\qc_workbook_{year}.xlsx"
    try:
        about = pd.read_excel(path, sheet_name="_About", header=None)
        row = about[about[0] == "Quarter"]
        if not row.empty:
            # Extract the integer form something like "Q2" or "Quarter: Q3"
            qn = int(re.search(r'\d+', str(row.iloc[0,1])).group())
            return max(1, min(4, qn))
    except Exception:
        pass
    return 4 

def months_for_q(q: int) -> List[str]:
    # Return month labels for a given quarter number
    return {1:['Jan','Feb','Mar'], 2:['Apr','May','Jun'], 3:['Jul','Aug','Sep'], 4:['Oct','Nov','Dec']}[int(q)]

def current_q_month_labels(current_year: int, reporting_q: int) -> List[str]:
    #Return labels like '2025-Feb' for months in the current reporting quarter
    return [f"{current_year}-{m}" for m in months_for_q(reporting_q)]

def qc_row_slice(df: pd.DataFrame, entity: str, wc: str, subq: str) -> pd.DataFrame:
    #Return the row for a (Entity, Worker Category, Subquestion)
    cond = (df[ENTITY_COL] == entity) & (df[WC_COL] == wc)
    if SUBQ_COL in df.columns and subq != "N/A":
        cond &= (df[SUBQ_COL] == subq)
    return df[cond]

# ================== VR loader & matching helpers (to fetch FI justifications) =====================
def _norm(s) -> str:
    # Normalize strings: trim, collapse spaces, lowercase
    return re.sub(r"\s+", " ", str(s).strip().lower())

@st.cache_data
def load_vr_variance(vr_path: str) -> pd.DataFrame | str:
    # Load the VR (staging) Excel and prepare helper columns for matching
    # Returns DataFrame if available, 'PENDING' if path is blank/missing, or error message on failure
    if not vr_path or not os.path.exists(vr_path):
        return "PENDING"  # special marker to show 'Pending submission'
    try:
        df = pd.read_excel(vr_path, sheet_name="Variance")
        needed = ["Entity Name","Year","Quarter","Month","Question","Subquestion","Worker Category","%Growth","Justification"]
        missing = [c for c in needed if c not in df.columns]
        if missing:
            return f"VR file missing columns: {missing}"
        
        # Normalize key matching columns
        df["_ent"] = df["Entity Name"].map(_norm)
        df["_subq"] = df["Subquestion"].map(_norm)
        df["_wc"] = df["Worker Category"].map(_norm)
        # Extract quarter number as numeric from a value like "Q1" "Q2"
        df["_qnum"] = df["Quarter"].astype(str).str.extract(r"\((\d)\)").astype(float)
        # Normalize month to Jan/Feb/... (first 3 letters)
        df["_month"]= df["Month"].astype(str).str[:3]
        return df
    except Exception as e:
        return f"Error reading VR file: {e}"

def _question_code_from_dataset(dataset_key: str) -> str:
    # Map Q1A: Employees -> Q1A for matching against VR's 'Question'
    return dataset_key.split(":")[0].replace(" ", "")

def find_vr_just_for_periods(
    vr_df: pd.DataFrame | str,
    dataset_key: str,
    entity_name: str,
    subq: str,
    wc: str,
    periods: List[str],   # e.g. ["2025-Feb"] or ["Q1 2025"]
) -> str:
    # Look up FI 'Justification' text in the VR file for the specified (entity/subquestion/worker category) and periods.
    # Accept either monthly labels "YYY-Mmm" or quarterly labels "Qx YYYY"
    # If multiple justifications match, deduplicate and join them with newlines.

    # Handle error / pending states cleanly
    if isinstance(vr_df, str):
        return "Pending submission" if vr_df == "PENDING" else vr_df
    
    qcode = _question_code_from_dataset(dataset_key)
    ent_norm = _norm(entity_name)
    subq_norm = _norm(subq)
    wc_norm = _norm(wc)
    justs = []

    for p in periods:
        # Monthly label "YYYY-Mmm"
        if "-" in p and "Q" not in p:
            yr_str, mon = p.split("-")
            try:
                yr = int(yr_str)
            except:
                continue
            # Strict month match
            sub = vr_df[
                (vr_df["Year"] == yr) &
                (vr_df["_month"] == mon) &
                (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                (vr_df["_ent"] == ent_norm) &
                (vr_df["_wc"] == wc_norm)
            ]
            if subq_norm and subq_norm != _norm("N/A"):
                sub = sub[sub["_subq"] == subq_norm]
            
            # Fallback: use quarter row if monthly row missing
            if sub.empty:
                q_from_mon = {"Jan":1,"Feb":1,"Mar":1,"Apr":2,"May":2,"Jun":2,
                              "Jul":3,"Aug":3,"Sep":3,"Oct":4,"Nov":4,"Dec":4}.get(mon, None)
                if q_from_mon is not None:
                    sub = vr_df[
                        (vr_df["Year"] == yr) &
                        (vr_df["_qnum"] == q_from_mon) &
                        (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                        (vr_df["_ent"] == ent_norm) &
                        (vr_df["_wc"] == wc_norm)
                    ]
                    if subq_norm and subq_norm != _norm("N/A"):
                        sub = sub[sub["_subq"] == subq_norm]
        else:
            # Quarterly label "Qx YYYY"
            try:
                qlab, yr_str = p.split()
                qn = int(qlab[1:])
                yr = int(yr_str)
            except:
                continue
            sub = vr_df[
                (vr_df["Year"] == yr) &
                (vr_df["_qnum"] == qn) &
                (vr_df["Question"].astype(str).str.upper() == qcode.upper()) &
                (vr_df["_ent"] == ent_norm) &
                (vr_df["_wc"] == wc_norm)
            ]
            if subq_norm and subq_norm != _norm("N/A"):
                sub = sub[sub["_subq"] == subq_norm]

        # Collect non-empty justifications and dedupe
        js = [j for j in sub["Justification"].astype(str).tolist()
              if str(j).strip() and str(j).strip().lower() != "nan"]
        if js:
            justs.append(" \n".join(sorted(set(js))))

    if not justs:
        return "—"
    return " \n".join(sorted(set(justs)))

# ===================== Multi-year Series Builders (monthly/quarterly) ========================
def build_multi_year_monthly_series(
    entity: str, wc: str, subq: str, sheet_key: str,
    start_year: int, end_year: int, excluded_years: set[int]
) -> Tuple[pd.Series, Optional[pd.Series]]:
    
    # Build a monthly time series from start_year..end_year (exlcuding excluded_years)
    # Index like 'YYY-Mmm'
    # Also returns a "YoY alignment series" (previous year's same month values aligned on current index) for YoY% calculations 
    vals, idx = [], []

    for yr in range(int(start_year), int(end_year)+1):
        if yr in excluded_years: 
            continue

        df = load_qc_sheet(yr, SHEET_MAP[sheet_key])
        if isinstance(df, str): 
            continue

        row = qc_row_slice(df, entity, wc, subq)
        if row.empty: 
            continue

        rq = get_reporting_quarter(yr)
        months = [m for m in ALL_MONTHS[:rq*3] if m in row.columns]
        if not months: 
            continue

        s = pd.to_numeric(row[months].iloc[0], errors="coerce").astype(float)

        for m in months:
            idx.append(f"{yr}-{m}")
            vals.append(s.get(m, np.nan))

    series = pd.Series(vals, index=idx, dtype=float)

    # Build YoY alignment for each YYYY-Mmm, pick value from (YYYY-1)-Mmm if available
    yoy_vals, yoy_idx = [], []

    for lab in series.index:
        yr, mon = lab.split('-')
        prev = f"{int(yr)-1}-{mon}"
        if prev in series.index and not pd.isna(series[prev]):
            yoy_idx.append(lab); yoy_vals.append(series[prev])

    yoy_series = pd.Series(yoy_vals, index=yoy_idx, dtype=float) if yoy_idx else None
    return series, yoy_series

def build_multi_year_quarterly_series(
    entity: str, wc: str, subq: str, sheet_key: str,
    start_year: int, end_year: int, excluded_years: set[int]
) -> Tuple[pd.Series, Optional[pd.Series]]:
    # Build a quarterly time series. Uses Qx_Total if present; otherwise sums the months of each quarter
    # Index like 'Qx YYYY'
    # Also constructs a YoY alignment series for YoY% calculations
    labels, values = [], []

    for yr in range(int(start_year), int(end_year)+1):
        if yr in excluded_years: 
            continue

        df = load_qc_sheet(yr, SHEET_MAP[sheet_key])
        if isinstance(df, str): 
            continue

        row = qc_row_slice(df, entity, wc, subq)
        if row.empty: 
            continue

        rq = get_reporting_quarter(yr)
        
        for q in range(1, rq+1):
            col = f"Q{q}_Total"
            if col in row.columns and not pd.isna(row.iloc[0][col]):
                v = float(pd.to_numeric(row.iloc[0][col], errors="coerce"))
            else:
                # Sum months if Qx_Total is missing
                mlist = [m for m in months_for_q(q) if m in row.columns]
                v = float(pd.to_numeric(row[mlist].iloc[0], errors="coerce").astype(float).sum()) if mlist else np.nan

            if not np.isnan(v):
                labels.append(f"Q{q} {yr}")
                values.append(v)

    series = pd.Series(values, index=labels, dtype=float)

    # Build YoY alignment for quarters
    yoy_vals, yoy_idx = [], []

    for lab in series.index:
        q, yr_str = lab.split()
        prev = f"{q} {int(yr_str)-1}"
        if prev in series.index and not pd.isna(series[prev]):
            yoy_idx.append(lab); yoy_vals.append(series[prev])

    yoy_series = pd.Series(yoy_vals, index=yoy_idx, dtype=float) if yoy_idx else None
    return series, yoy_series

# ===================== Outlier detection engine ====================

def find_outliers_v2(
    series: pd.Series,
    yoy_series: Optional[pd.Series],
    pct_thresh: float,   # MoM/QoQ % threshold (gate)
    abs_cutoff: float,   # absolute change threshold (significance)
    iqr_k: float,        # IQR multiplier (significance)
    yoy_thresh: float    # YoY % threshold (significance)
) -> pd.DataFrame:
    
    # Outlier logic (two-stage):
    # 1) Gate: High MoM%/QoQ% if |MoM| >= pct_thresh AND |Δ| >= abs_cutoff
    # 2) Only then flag if (YoY >= yoy_thresh OR value is IQR outlier)
    # Reasons include: "High MoM%", "High YoY%", "IQR Detect Outlier"
    # Returns a DataFrame with index=Period and columns: Value, Statistical Reasons.
    
    out = []
    clean = series.dropna()
    if clean.size < 2:
        return pd.DataFrame(columns=["Period", "Value", "Statistical Reasons"]).set_index("Period")
    
    # IQR bounds
    q1, q3 = clean.quantile(0.25), clean.quantile(0.75)
    iqr = q3 - q1
    lb, ub = ((q1 - iqr_k * iqr, q3 + iqr_k * iqr) if iqr > 0 else (None, None))

    for i, (period, cur) in enumerate(series.items()):
        if pd.isna(cur) or i == 0: 
            continue
        prev = series.iloc[i - 1]
        if pd.isna(prev) or prev == 0: 
            continue

        abs_chg = cur - prev
        mom = abs_chg / prev

        # Stage 1: Gate on MoM/QOQ % AND absolute change
        if abs(mom) >= pct_thresh and abs(abs_chg) >= abs_cutoff:
            reasons = [f"High MoM% ({mom:+.0%})"]
            extra_reason = False

            # IQR anomaly on level
            if iqr > 0 and (cur < lb or cur > ub):
                reasons.append("IQR Detect Outlier")
                extra_reason = True

            # YoY significance if aligned previous year's period exists
            if yoy_series is not None and period in yoy_series.index:
                py = yoy_series.get(period)
                if not pd.isna(py) and py != 0:
                    yoy = (cur - py) / py
                    if abs(yoy) >= yoy_thresh:
                        reasons.append(f"High YoY% ({yoy:+.0%})")
                        extra_reason = True

            # Only flag if at least one significance reason exists
            if extra_reason:
                out.append({
                    "Period": period,
                    "Value": f"{cur:,.2f}",
                    "Statistical Reasons": ", ".join(reasons)
                })
    return pd.DataFrame(out).set_index("Period") if out else pd.DataFrame(columns=["Period", "Value", "Statistical Reasons"]).set_index("Period")

# ============= Charting: dual axis bar (levels) + line (% change) + outlier markers
def plot_dual_axis_with_outliers(
    series: pd.Series,
    growth_pct: pd.Series,
    outliers_focus: pd.DataFrame,
    title: str,
    left_title: str = "Value",
    right_title: str = "% Change (MoM)"
):
    # Render a Plotly Chart:
    # - Left axis (bars): level values
    # - Right axis (line): % change vs previous period
    # - Red X markers for 'true outliers' within the focus window

    x = list(series.index)
    y = series.values.astype(float)
    g = growth_pct.reindex(series.index).fillna(0)

    fig = go.Figure()

    # Bars for values (left axis)
    fig.add_trace(go.Bar(
        x=x, y=y, name=left_title, yaxis="y1", opacity=0.75,
        hovertemplate='Period: %{x}<br>Value: %{y:,.0f}'
    ))

    # Line for % change (right axis)
    fig.add_trace(go.Scatter(
        x=x, y=g.values, name=right_title, yaxis="y2",
        mode="lines+markers", line=dict(width=2, dash="dot", color="#003366"),
        hovertemplate='Period: %{x}<br>% Growth: %{y:+.0f}%'
    ))
    # Outlier markers (red X)
    if not outliers_focus.empty:
        ox = [p for p in outliers_focus.index if p in series.index]
        oy = [float(g.loc[p]) for p in ox]
        oreason = [outliers_focus.loc[p, "Statistical Reasons"] for p in ox]
        fig.add_trace(go.Scatter(
            x=ox, y=oy, mode='markers', name='True Outlier', yaxis="y2",
            marker=dict(symbol='x', size=14, color='red', line_width=2),
            hovertemplate='Period: %{x}<br>% Growth: %{y:+.0f}%<br>%{customdata}',
            customdata=oreason
        ))

    fig.update_layout(
        title=dict(text=title, font=dict(size=16), x=0.5, xanchor='center'),
        hovermode='x unified',
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='left', x=0),
        margin=dict(l=10, r=10, t=70, b=10),
        xaxis=dict(tickangle=45),
        yaxis=dict(title=left_title, side='left'),
        yaxis2=dict(title=right_title, overlaying='y', side='right', showgrid=False, tickformat=".0f")
    )
    st.plotly_chart(fig, use_container_width=True)

# ==================== Hierarchy Loader (Parent-Child OR Institution/Sector/Subsector)
@st.cache_data
def load_hierarchy_any(path: str) -> Dict[str, List[str]]:
    
    # Accepts either:
    #  - Parent,Child CSV (used directly), or
    #  - Institution,Sector,Subsector CSV (auto-builds AFI tree):
    #       AFI -> Banking Institution / DFI / Insurans-Takaful
    #       Banking Institution -> (Commercial, Investment, Foreign, Islamic, Digital, IIB) -> entities
    #       DFI -> entities directly
    #       Insurans/Takaful -> (Insurers, Takaful Operators) -> entities
    # Removes self-child edges.

    import pandas as pd, os
    if not path or not os.path.exists(path):
        return {}

    df = pd.read_csv(path)
    df.columns = [c.strip() for c in df.columns]

    # Case A: Parent-Child edges provided directly
    if {"Parent","Child"} <= set(df.columns):
        df["Parent"] = df["Parent"].astype(str).str.strip()
        df["Child"]  = df["Child"].astype(str).str.strip()
        df = df[df["Parent"].str.lower() != df["Child"].str.lower()]
        return df.groupby("Parent")["Child"].apply(list).to_dict()

    # Case B: Institution,Sector,Subsector (build All FI tree)
    if {"Institution","Sector","Subsector"} <= set(df.columns):
        AFI  = "All Financial Institutions"
        BI   = "Banking Institution"
        DFI  = "DFI"
        ITOP = "Insurans/Takaful Operators"

        df["Institution"] = df["Institution"].astype(str).str.strip()
        df["Sector"]      = df["Sector"].astype(str).str.strip()
        df["Subsector"]   = df["Subsector"].astype(str).str.strip()

        banking_subs = {"Commercial Banks","Investment Banks","Foreign Banks","Islamic Banks","Digital Banks","International Islamic Banks"}
        ins_tak_subs = {"Insurers","Takaful Operators"}
        present_subs = set(df["Subsector"].unique())

        edges = []

        # All FI layer
        if (banking_subs & present_subs): edges.append({"Parent": AFI, "Child": BI})
        if ("DFI" in present_subs):      edges.append({"Parent": AFI, "Child": DFI})
        if (ins_tak_subs & present_subs):edges.append({"Parent": AFI, "Child": ITOP})

        # Banking branches
        for s in sorted(banking_subs & present_subs):
            edges.append({"Parent": BI, "Child": s})
        # Ins/Takaful branches
        for s in sorted(ins_tak_subs & present_subs):
            edges.append({"Parent": ITOP, "Child": s})

        # DFI -> leaf institutions (NO subsector under DFI)
        if "DFI" in present_subs:
            for inst in sorted(df.loc[df["Subsector"] == "DFI", "Institution"].unique()):
                edges.append({"Parent": DFI, "Child": inst})

        # Banking leaves
        for s in sorted(banking_subs & present_subs):
            for inst in sorted(df.loc[df["Subsector"] == s, "Institution"].unique()):
                edges.append({"Parent": s, "Child": inst})

        # Ins/Takaful leaves
        for s in sorted(ins_tak_subs & present_subs):
            for inst in sorted(df.loc[df["Subsector"] == s, "Institution"].unique()):
                edges.append({"Parent": s, "Child": inst})

        hier_df = pd.DataFrame(edges).drop_duplicates()
        hier_df = hier_df[hier_df["Parent"].str.lower() != hier_df["Child"].str.lower()]
        return hier_df.groupby("Parent")["Child"].apply(list).to_dict()

    # If neither schema matches, warn and return empty
    st.warning("Hierarchy CSV must have either (Parent,Child) or (Institution,Sector,Subsector) columns.")
    return {}

# ================ Attribution helpers: discover leaf entities present in QC =======================

def _norm2(s: str) -> str:
    # Normalize but keep casefold (better for non-english) and preserve space a single spaces
    return re.sub(r"\s+", " ", str(s).strip()).casefold()

def _canonical_map(series: pd.Series) -> Dict[str, str]:
    # Build a map 'normalized name' -> 'most frequent display variant' observed in QC.
    # This avoids display mismatches when the same entity appears with minor naming differences.
    tmp = series.dropna().astype(str)
    norm = tmp.map(_norm2)
    canon_map = (
        pd.DataFrame({"orig": tmp, "norm": norm})
        .groupby("norm")["orig"]
        .agg(lambda s: s.value_counts().idxmax())
        .to_dict()
    )
    return canon_map

def get_children_in_data(df: pd.DataFrame, parent: str, hier: Dict[str, List[str]]) -> List[str]:
    
    # Return immediate children of 'parent' that actually appear in QC (or exist as groups to recurse into).
    # - Ignores self-child edges.
    # - If a child group exists in the hierarchy but has no direct QC row, we keep it to recurse down later.
    if parent not in hier:
        return []
    
    candidates = [c for c in hier[parent] if _norm2(c) != _norm2(parent)]

    # Build canonical map for QC entities to ensure display names match QC variants
    canon = _canonical_map(df[ENTITY_COL])
    present_norms = set(canon.keys())

    matched = []
    for c in candidates:
        n = _norm2(c)
        if n in present_norms:
            matched.append(canon[n])     # exact QC display name
        elif c in hier:
            matched.append(c)            # group exists in hierarchy (no direct QC row)
    return list(dict.fromkeys(matched))

def flatten_leaves_in_qc(df: pd.DataFrame, root: str, hier: Dict[str, List[str]]) -> List[str]:
    
    # Recursively collect leaf entities under 'root' that actually exist in QC.
    # A leaf is a child that is not a parent in 'hier' (i.e., no further children).
    
    leaves = set()

    def _recurse(node: str):
        kids = get_children_in_data(df, node, hier)
        if not kids:
            #If node is a real entity row in QC, mark as leaf
            if node in df[ENTITY_COL].unique():
                leaves.add(node)
            return
        for k in kids:
            if k in hier:
                _recurse(k)
            else:
                leaves.add(k)

    _recurse(root)

    # Remove known group labels if they slipped in
    groups_to_exclude = {
        "All Financial Institutions", "Banking Institution", "DFI",
        "Insurans/Takaful Operators", "Commercial Banks", "Investment Banks",
        "Foreign Banks", "Islamic Banks", "Digital Banks", "International Islamic Banks",
        "Insurers", "Takaful Operators"
    }
    return sorted([e for e in leaves if e not in groups_to_exclude])

def compute_entity_contributions(
    entities: List[str],
    subq: str,
    wc: str,
    sheet_key: str,
    start_year: int,
    end_year: int,
    exclude_years: set[int],
    period_label: str,     # "YYYY-Mmm" or "Qx YYYY"
    time_view: str         # "Monthly" or "Quarterly"
) -> pd.DataFrame:
    
    # For a chosen outlier period, compute per-entity:
    #  - Prev (previous period value), Curr (current period), Δ (Curr-Prev)
    #  - Contribution % relative to the total Δ across entities
    # Sorted by |Δ| descending to surface top movers.

    rows = []

    for ent in entities:
        if time_view == "Monthly":
            s, _ = build_multi_year_monthly_series(ent, wc, subq, sheet_key, start_year, end_year, exclude_years)
        else:
            s, _ = build_multi_year_quarterly_series(ent, wc, subq, sheet_key, start_year, end_year, exclude_years)

        if s.empty or period_label not in s.index:
            continue
        idx_list = list(s.index)
        i = idx_list.index(period_label)
        if i == 0:
            continue  # no previous period
        prev, cur = float(s.iloc[i-1]), float(s.iloc[i])
        if np.isnan(prev) or np.isnan(cur):
            continue
        rows.append({"Entity": ent, "Prev": prev, "Curr": cur, "Delta": cur - prev})

    dfe = pd.DataFrame(rows)
    if dfe.empty:
        return dfe

    total_delta = dfe["Delta"].sum()
    dfe["Contribution %"] = np.where(total_delta != 0, dfe["Delta"] / total_delta * 100.0, np.nan)

    # Sort by magnitude of movement
    dfe.sort_values(by="Delta", key=np.abs, ascending=False, inplace=True)  # rank biggest movers
    dfe.reset_index(drop=True, inplace=True)
    return dfe

def plot_top_contributors_bar(dfe: pd.DataFrame, title: str = "Top Contributors (Δ vs previous)"):
    # Horizontal bar chart highlighting which entities contributed most 
    # Positive = increase, Negative = decrease
    if dfe.empty:
        st.info("No entity-level data available.")
        return
    
    show = dfe.copy()
    show["sign"] = np.where(show["Delta"] >= 0, "Increase", "Decrease")
    colors = {"Increase": "#2ca02c", "Decrease": "#d62728"}

    fig = go.Figure(go.Bar(
        x=show["Delta"],
        y=show["Entity"],
        orientation="h",
        marker=dict(color=[colors[s] for s in show["sign"]]),
        hovertemplate="Entity: %{y}<br>Δ: %{x:+,.0f}<extra></extra>"
    ))

    fig.update_layout(
        title=title,
        xaxis_title="Δ vs previous period",
        yaxis_title=None,
        margin=dict(l=10, r=10, t=50, b=10)
    )

    st.plotly_chart(fig, use_container_width=True)


# ==================== Sidebar UI Controls ==================

def sidebar_controls():
    st.sidebar.title("Analysis Controls")

    # Years / timeline selection
    years_available = list(range(2019, 2031))
    current_year = st.sidebar.selectbox(
        "Current Year (detect current quarter from this workbook):",
        options=sorted(years_available, reverse=True),
        index=years_available.index(2025)
    )
    start_year = st.sidebar.selectbox(
        "Start Year (timeline):",
        options=years_available,
        index=years_available.index(2022)
    )
    end_year = st.sidebar.selectbox(
        "End Year (timeline):",
        options=years_available,
        index=years_available.index(2025)
    )
    if end_year < start_year:
        st.sidebar.error("End Year must be ≥ Start Year.")
    exclude_years = st.sidebar.multiselect(
        "Exclude Years (optional):",
        options=[y for y in range(start_year, end_year+1)],
        default=[]
    )

    # Threshold for outlier detection
    st.sidebar.markdown("---")
    st.sidebar.subheader("Thresholds")
    mom_pct = st.sidebar.slider("MoM/QoQ % Threshold (Gate)", 0, 100, 25, 5, format="%d%%") / 100.0
    abs_cut = st.sidebar.slider("Absolute Change (Significance)", 10, 1000, 50, 10)
    iqr_k = st.sidebar.slider("IQR Sensitivity (Significance)", 1.0, 3.0, 1.5, 0.1)
    yoy_pct = st.sidebar.slider("YoY % Threshold (Significance)", 0, 100, 30, 5, format="%d%%") / 100.0

    # Frequency & dataset
    st.sidebar.markdown("---")
    time_view = st.sidebar.radio("Frequency:", options=['Monthly','Quarterly'], horizontal=True)
    dataset = st.sidebar.selectbox("Dataset:", options=list(SHEET_MAP.keys()))

    # Outlier focus window selection
    st.sidebar.markdown("---")
    st.sidebar.subheader("Outlier Focus")
    focus_mode = st.sidebar.radio(
        "Show outliers for:",
        options=["Current quarter", "Pick year & quarter"],
        index=0, horizontal=False
    )
    focus_year = None
    focus_quarter = None
    if focus_mode == "Pick year & quarter":
        focus_year = st.sidebar.selectbox(
            "Focus Year:",
            options=list(range(start_year, end_year + 1)),
            index=list(range(start_year, end_year + 1)).index(min(end_year, int(current_year)))
        )
        focus_quarter = st.sidebar.selectbox("Focus Quarter:", options=[1, 2, 3, 4], index=0)

    # VR staging file path
    st.sidebar.markdown("---")
    st.sidebar.subheader("VR Staging")
    vr_path = st.sidebar.text_input(
        "Full path to VR staging Excel (sheet 'Variance'):",
        value="",
        placeholder=r"C:\...\VR_Consol_2025_Quarter1.xlsx"
    )

    # Hierarchy CSV path 
    st.sidebar.markdown("---")
    st.sidebar.subheader("Hierarchy (optional)")
    hier_path = st.sidebar.text_input(
        "Path to Entity Hierarchy CSV (Parent,Child or Institution,Sector,Subsector):",
        value="",
        placeholder=r"C:\...\entity_hierarchy_full.csv"
    )

    return {
        "current_year": int(current_year),
        "start_year": int(start_year),
        "end_year": int(end_year),
        "exclude_years": set(map(int, exclude_years)),
        "mom_pct": float(mom_pct),
        "abs_cut": float(abs_cut),
        "iqr_k": float(iqr_k),
        "yoy_pct": float(yoy_pct),
        "time_view": time_view,
        "dataset": dataset,
        "focus_mode": focus_mode,
        "focus_year": int(focus_year) if focus_year is not None else None,
        "focus_quarter": int(focus_quarter) if focus_quarter is not None else None,
        "vr_path": vr_path.strip(),
        "hier_path": hier_path.strip(),
    }

# ==================== Main Interactive View ==================

def _is_total_wc(name: str) -> bool:
    # Consider anything that contains 'total' (case-insensitive) as a Total WC
    return "total" in str(name).casefold()

# ==========================================
# VR helper: Build VR table by Worker Category (non-destructive addition)
# ==========================================
def build_vr_wc_table_for_entity(vr_df: pd.DataFrame, df_cur: pd.DataFrame, dataset_key: str, entity: str, subq: str, periods_list: List[str]) -> pd.DataFrame:
    """
    Returns a dataframe of VR justifications by Worker Category for the given entity/subquestion and period list.
    If vr_df is "PENDING" or an error string, returns a single-row dataframe.
    """
    if isinstance(vr_df, str):
        return pd.DataFrame([{"Worker Category": "All workers", "FI Justification (selected period)": ("Pending submission" if vr_df == "PENDING" else vr_df)}])

    try:
        mask_ent = (df_cur[ENTITY_COL] == entity)
        if SUBQ_COL in df_cur.columns:
            mask_ent &= (df_cur[SUBQ_COL] == subq)
        wc_list = sorted(df_cur.loc[mask_ent, WC_COL].dropna().unique().tolist())
        if not wc_list:
            wc_list = ["All workers"]
    except Exception:
        wc_list = ["All workers"]

    rows = []
    for wc_any in wc_list:
        try:
            just_text = find_vr_just_for_periods(
                vr_df=vr_df,
                dataset_key=dataset_key,
                entity_name=entity,
                subq=subq,
                wc=wc_any,
                periods=periods_list
            )
        except Exception as e:
            just_text = f"(error reading VR: {e})"
        rows.append({"Worker Category": wc_any, "FI Justification (selected period)": just_text})

    show_wc = pd.DataFrame(rows)
    if not show_wc.empty:
        # Sort with "All workers" (or total-like) first, then alphabetical
        show_wc["__is_total"] = show_wc["Worker Category"].map(lambda x: 0 if _is_total_wc(str(x)) else 1)
        show_wc.sort_values(["__is_total", "Worker Category"], inplace=True, kind="stable")
        show_wc.drop(columns="__is_total", inplace=True)
    return show_wc


def main_view():
    st.title("🏦 LMS Analysis Dashboard")

    # Read sidebar configuration and show a summary 
    cfg = sidebar_controls()
    rq = get_reporting_quarter(cfg["current_year"])
    st.sidebar.info(f"Analyzing **{cfg['start_year']}–{cfg['end_year']}** (excl: {', '.join(map(str,cfg['exclude_years'])) or 'none'}) • Current workbook: **{cfg['current_year']} Q{rq}**")

    # Decide which period's outliers to highlight (current quarter or user-picked)
    if cfg["focus_mode"] == "Current quarter":
        focus_q = rq
        focus_year = cfg["current_year"]
    else:
        focus_q = cfg["focus_quarter"] or rq
        focus_year = cfg["focus_year"] or cfg["current_year"]

    focus_month_labels = set([f"{focus_year}-{m}" for m in months_for_q(focus_q)])
    focus_quarter_label = f"Q{focus_q} {focus_year}"

    # Badge display
    month_str = "–".join(months_for_q(focus_q))
    st.markdown(
        f'Outlier focus: **Q{focus_q} {focus_year}** ({month_str})',
        unsafe_allow_html=True
    )

    # Load the current-year QC sheet for the choosen dataset
    df_cur = load_qc_sheet(cfg["current_year"], SHEET_MAP[cfg["dataset"]])
    if isinstance(df_cur, str):
        st.error(df_cur)
        return
    


    # Normalize display variations for Entity/Subquestion/Worker Category to keep UI clean
    def _norm_local(s: str) -> str:
        return re.sub(r"\s+", " ", str(s).strip()).casefold()
    
    def _canonical(series: pd.Series) -> pd.Series:

        # Convert each value to its most frequent 'display' variant based on its normalized form
        tmp = series.dropna().astype(str)
        norm = tmp.map(_norm_local)
        canon_map = (
            pd.DataFrame({"orig": tmp, "norm": norm})
            .groupby("norm")["orig"]
            .agg(lambda s: s.value_counts().idxmax())
        )
        return norm.map(canon_map)

    df_cur["_ent_norm"] = df_cur[ENTITY_COL].map(_norm_local)
    df_cur["_ent_disp"] = _canonical(df_cur[ENTITY_COL])
    if SUBQ_COL in df_cur.columns:
        df_cur["_subq_norm"] = df_cur[SUBQ_COL].map(_norm_local)
        df_cur["_subq_disp"] = _canonical(df_cur[SUBQ_COL])
    df_cur["_wc_norm"] = df_cur[WC_COL].map(_norm_local)
    df_cur["_wc_disp"] = _canonical(df_cur[WC_COL])

    # Entity/Subquestion/Worker Category selection widget
    st.header(f"Outlier Detection: {cfg['dataset']}")
    entity = st.selectbox(
        "Entity / Group:",
        options=df_cur[ENTITY_COL].unique(),
        index=list(df_cur[ENTITY_COL].unique()).index(ROLLUP_KEY) if ROLLUP_KEY in df_cur[ENTITY_COL].unique() else 0
    )
    if SUBQ_COL in df_cur.columns and df_cur[df_cur[ENTITY_COL]==entity][SUBQ_COL].nunique() > 1:
        u_subq = df_cur[df_cur[ENTITY_COL]==entity][SUBQ_COL].unique()
        subq_default = np.where(u_subq == 'Employment = A+B(i)+B(ii)')[0][0] if 'Employment = A+B(i)+B(ii)' in u_subq else 0
        subq = st.selectbox("Subquestion:", options=u_subq, index=int(subq_default))
        u_wc = df_cur[(df_cur[ENTITY_COL]==entity) & (df_cur[SUBQ_COL]==subq)][WC_COL].unique()
        wc_default = np.where(u_wc == 'Total Employment')[0][0] if 'Total Employment' in u_wc else 0
        wc = st.selectbox("Worker Category:", options=u_wc, index=int(wc_default))
    else:
        subq = "N/A"
        u_wc = df_cur[df_cur[ENTITY_COL]==entity][WC_COL].unique()
        wc_default = np.where(u_wc == 'Total Employment')[0][0] if 'Total Employment' in u_wc else 0
        wc = st.selectbox("Worker Category:", options=u_wc, index=int(wc_default))

    st.caption(f"Displaying: {entity}  \n{subq}  \n{wc}")

    # Load VR staging once
    # String indicates "PENDING" or error
    vr_df = load_vr_variance(cfg["vr_path"])

    # Build the selected frequency series and run detection
    if cfg["time_view"] == "Monthly":
        series, yoy_series = build_multi_year_monthly_series(
            entity, wc, subq, cfg["dataset"],
            cfg["start_year"], cfg["end_year"],
            cfg["exclude_years"]
        )
        if series.empty or series.isnull().all():
            st.warning("No data found for the selected timeline/filters.")
            return
        
        # Detection outliers across the whole timeline
        out_all = find_outliers_v2(series, yoy_series, cfg["mom_pct"], cfg["abs_cut"], cfg["iqr_k"], cfg["yoy_pct"])

        # Keep only outliers within focused quarter's months
        out_focus = out_all.loc[out_all.index.intersection(focus_month_labels)]

        # Attach VR justifications per outlier period
        if not out_focus.empty:
            periods_list = out_focus.index.tolist()
            per_row_just = []
            if isinstance(vr_df, str):
                per_row_just = [("Pending submission" if vr_df == "PENDING" else vr_df)] * len(out_focus)
            else:
                for p in periods_list:
                    per_row_just.append(find_vr_just_for_periods(vr_df, cfg["dataset"], entity, subq, wc, [p]))

            out_focus = out_focus.copy()
            out_focus["FI Justification"] = per_row_just

        # %MoM growth for plotting (rounded to match VR)
        growth_pct = series.pct_change()
        growth_pct = growth_pct.replace([np.inf, -np.inf], np.nan)
        growth_pct = (growth_pct * 100).round().fillna(0)



        # Plot monthly chart
        plot_dual_axis_with_outliers(
            series=series, growth_pct=growth_pct, outliers_focus=out_focus,
            title=f"Monthly Trend ({cfg['start_year']}–{cfg['end_year']})"
        )

    else:
        # Quarterly path
        series, yoy_series = build_multi_year_quarterly_series(
            entity, wc, subq, cfg["dataset"],
            cfg["start_year"], cfg["end_year"],
            cfg["exclude_years"]
        )
        if series.empty or series.isnull().all():
            st.warning("No data found for the selected timeline/filters.")
            return
        
        # Detect outliers across the whole timeline
        out_all = find_outliers_v2(series, yoy_series, cfg["mom_pct"], cfg["abs_cut"], cfg["iqr_k"], cfg["yoy_pct"])

        # Keep only the focused quarter outlier
        out_focus = out_all.loc[out_all.index.intersection({focus_quarter_label})]

        # Attach VR justifications combining all months in that quarter
        if not out_focus.empty:
            qy = int(focus_quarter_label.split()[1])
            qn = int(focus_quarter_label.split()[0][1:])
            q_months = [f"{qy}-{m}" for m in months_for_q(qn)]
            per_row_just = []
            for _ in out_focus.index.tolist():  # usually one row
                per_row_just.append(
                    find_vr_just_for_periods(
                        vr_df=vr_df, dataset_key=cfg["dataset"], entity_name=entity,
                        subq=subq, wc=wc, periods=q_months
                    )
                )
            out_focus = out_focus.copy()
            out_focus["FI Justification"] = per_row_just

        # %QoQ growth for plotting
        growth_pct = series.pct_change()
        growth_pct = growth_pct.replace([np.inf, -np.inf], np.nan)
        growth_pct = (growth_pct * 100).round().fillna(0)

        # Plot quarterly chart
        plot_dual_axis_with_outliers(
            series=series, growth_pct=growth_pct, outliers_focus=out_focus,
            title=f"Quarterly Trend ({cfg['start_year']}–{cfg['end_year']})",
            right_title="% Change (QoQ)"
        )

    # KPI cards + outlier table for current focus
    c1,c2,c3,c4 = st.columns(4)
    c1.metric("Latest Value", f"{series.iloc[-1]:,.0f}")
    c2.metric("Average", f"{series.mean():,.0f}")
    c3.metric("Highest Value", f"{np.nanmax(series):,.0f}")
    c4.metric("Current-Q Outliers", len(out_focus))
    if not out_focus.empty:
        st.error("🚨 True Outlier(s) in Current Quarter")
        st.dataframe(out_focus, use_container_width=True)
    else:
        st.success("✅ No current-quarter outliers")

    # Attribution: which leaf entities contributed to the outlier?
    # Only shown when (a) there is at least one outlier in focus
    # (b) selected entity is the root roll-up "All Financial Institution"
    hier = load_hierarchy_any(cfg.get("hier_path","")) or HIERARCHY_MAP

    if not out_focus.empty and entity == ROLLUP_KEY:
        with st.expander("🔎 Contribution by FI", expanded=True):
            # Let the user pick which flagged period (within focus) to attribute
            attrib_period = st.selectbox(
                "Outlier period:",
                options=list(out_focus.index),
                index=0
            )

            # Resolve all leaves under All FI that actually exist in QC
            leaf_entities = flatten_leaves_in_qc(df_cur, ROLLUP_KEY, hier)
            if not leaf_entities:
                st.warning("No leaf entities found under All FI. Check the hierarchy path or mapping.")
            else:
                # Calculate contributions (Prev, Curr, Delta, Contributions %) per leaf for that period
                dfe = compute_entity_contributions(
                    entities=leaf_entities,
                    subq=subq,
                    wc=wc,
                    sheet_key=cfg["dataset"],
                    start_year=cfg["start_year"],
                    end_year=cfg["end_year"],
                    exclude_years=cfg["exclude_years"],
                    period_label=attrib_period,
                    time_view=cfg["time_view"]
                )

                if dfe.empty:
                    st.info("No entity had usable values in the selected period (or previous period).")
                else:
                    # Quick filtering and top-N selection
                    col_s, col_n = st.columns([2,1])
                    search_text = col_s.text_input("Filter entities (optional):", value="", placeholder="Type part of a name, e.g., 'MAYBANK'")
                    top_n = col_n.slider("Top N by |Δ|", min_value=5, max_value=50, value=15, step=5)

                    dfe_view = dfe.copy()
                    if search_text.strip():
                        q = search_text.strip().lower()
                        dfe_view = dfe_view[dfe_view["Entity"].str.lower().str.contains(q)]
                    dfe_view = dfe_view.head(top_n)

                    # Attach VR justification per entity for the same period
                    if isinstance(vr_df, str):
                        dfe_view["FI Justification"] = ("Pending submission" if vr_df == "PENDING" else vr_df)
                    else:
                        j_list = []
                        for ent in dfe_view["Entity"]:
                            j_list.append(
                                find_vr_just_for_periods(
                                    vr_df=vr_df,
                                    dataset_key=cfg["dataset"],
                                    entity_name=ent,
                                    subq=subq,
                                    wc=wc,
                                    periods=[attrib_period]
                                )
                            )
                        dfe_view["FI Justification"] = j_list



                    # --- Add-on: Interactive "FI Justification" click-to-expand (by Worker Category)
                    try:
                        show2 = dfe_view.copy()
                        show2["Prev"] = show2["Prev"].map(lambda v: f"{v:,.0f}")
                        show2["Curr"] = show2["Curr"].map(lambda v: f"{v:,.0f}")
                        show2["Delta"] = show2["Delta"].map(lambda v: f"{v:+,.0f}")
                        show2["Contribution %"] = show2["Contribution %"].map(lambda p: f"{p:+.1f}%" if pd.notna(p) else "–")
                        # Render a data editor with a clickable checkbox column
                        show2["FI Justification (all WC)"] = False

                        edited = st.data_editor(
                            show2,
                            use_container_width=True,
                            disabled=["Prev","Curr","Delta","Contribution %","FI Justification"],
                            column_config={
                                "FI Justification (all WC)": st.column_config.CheckboxColumn(
                                    "FI Justification", help="Click to view all Worker Categories for this FI"
                                )
                            },
                            key=f"fi_contrib_allwc_{attrib_period}"
                        )

                        # For any rows where the user ticked the FI Justification cell, render a breakdown by Worker Category
                        sel = edited[edited["FI Justification (all WC)"] == True]
                        if not sel.empty:
                            st.markdown("### VR Justifications by Worker Category (Selected FI)")
                            for _, _row in sel.iterrows():
                                _ent = _row["Entity"]
                                st.markdown(f"**{_ent} — {attrib_period}**")
                                wc_table = build_vr_wc_table_for_entity(
                                    vr_df=vr_df,
                                    df_cur=df_cur,
                                    dataset_key=cfg["dataset"],
                                    entity=_ent,
                                    subq=subq,
                                    periods_list=[attrib_period]
                                )
                                st.dataframe(wc_table, use_container_width=True)
                    except Exception as _e:
                        st.caption(f"(FYI: could not render interactive VR-by-worker-category view: {_e})")


                    # Visualize contributions
                    plot_top_contributors_bar(dfe_view, title=f"Top Contributors (All FI — {attrib_period})")
    else:
        # Keep UI minimal if not applicable
        pass


# ============ Run ================
if __name__ == "__main__":
    router()


# ----------------------------
# Page Navigation (Added)
# ----------------------------
def router():
    page = st.sidebar.selectbox(
        "Select Page",
        ["Dashboard", "MHS Table"]
    )
    if page == "Dashboard":
        main_view()
    elif page == "MHS Table":
        mhs_page()


